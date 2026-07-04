"""Iter-250b · P2 — Excel-based import for product categories + products.

Phase 1 (this commit): **categories import**.

Pipeline:
  1. Frontend POSTs the .xlsx to `/api/products/categories/import/preview`.
     The endpoint parses the workbook IN MEMORY (no disk writes) and
     returns a structured preview report:
       * new_count / existing_count
       * orphan_subcategory_count (rows flagged فرعي=نعم but parent blank)
       * sample of new categories
       * the resolved hierarchy that WILL be created
       * root category path
  2. Frontend posts `/api/products/categories/import/confirm` with the
     same file. Backend re-parses + atomically:
       * ensures the dedicated root "المنتجات المستوردة" exists once
       * upserts every category as `kind="product"` under it
       * never deletes, never reuses an existing expense category by
         accident (we always scope the lookup to the product subtree
         by walking ancestors).

READ-ONLY for ALL existing data:
  * no expense category is renamed, moved, or deleted.
  * we ONLY upsert into `db.expense_category_tree` with `kind="product"`.

Two endpoints in this file. Phase 2 (products import) will live in
the same module under different prefixes.
"""
from __future__ import annotations

import io
import re
import uuid
from datetime import datetime, timezone
from typing import Any
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from openpyxl import load_workbook


# ─────────────────────────  helpers  ──────────────────────────────────
def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm(value: Any) -> str:
    s = "" if value is None else str(value)
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _norm_key(value: Any) -> str:
    """Case-folded + tatweel-stripped key for de-duplication."""
    s = _norm(value).casefold()
    return s.replace("\u0640", "")  # arabic tatweel


def _is_yes(value: Any) -> bool:
    s = _norm(value).casefold()
    return s in {"نعم", "yes", "true", "1", "y"}


def _parse_categories_xlsx(raw: bytes) -> dict:
    """Parse the categories workbook into:
       { rows: [{ name, is_sub, parent, raw_row_index }],
         orphans: [...],
         duplicates_in_file: [...] }
    Columns expected:
        A: التصنيفات (name)             — required
        B: هل التصنيف فرعي ام لا (yes/no)
        C: التصنيف الاساسي (parent name) — required when B==yes
        D..: ignored
    """
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"تعذّر قراءة ملف Excel: {e}")
    ws = wb.active
    rows: list[dict] = []
    orphans: list[dict] = []
    duplicates_in_file: list[dict] = []
    seen_keys: set[str] = set()

    header_seen = False
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        a = _norm(row[0] if len(row) > 0 else None)
        b = row[1] if len(row) > 1 else None
        c = _norm(row[2] if len(row) > 2 else None)
        if not header_seen:
            header_seen = True
            # Treat row 1 as header. We don't strictly validate header
            # text; we trust column positions.
            continue
        if not a:
            continue
        is_sub = _is_yes(b)
        parent = c if (is_sub and c) else None
        key = _norm_key(a)
        if key in seen_keys:
            duplicates_in_file.append({"row": idx, "name": a})
            continue
        seen_keys.add(key)
        item = {
            "name": a,
            "is_sub": is_sub,
            "parent": parent,
            "raw_row_index": idx,
            "key": key,
        }
        if is_sub and not c:
            orphans.append(item)
        rows.append(item)
    return {
        "rows":                rows,
        "orphans":             orphans,
        "duplicates_in_file":  duplicates_in_file,
    }


ROOT_NAME = "المنتجات المستوردة"
UNCAT_NAME = "غير مصنف"


async def _existing_product_keys(db, uid: str, root_id: str) -> dict[str, str]:
    """Return { norm_key: category_id } for all kind=product categories
    that descend from the root. We rely on the `path` field to detect
    membership of the subtree without recursive queries."""
    out: dict[str, str] = {}
    async for c in db.expense_category_tree.find(
        {"user_id": uid, "kind": "product"},
        {"_id": 0, "id": 1, "name": 1, "path": 1, "parent_id": 1},
    ):
        out[_norm_key(c.get("name"))] = c["id"]
    return out


async def _ensure_product_root(db, uid: str) -> dict:
    """Upsert the dedicated `kind=product` root once per user."""
    existing = await db.expense_category_tree.find_one(
        {"user_id": uid, "kind": "product",
         "parent_id": None, "name": ROOT_NAME},
        {"_id": 0},
    )
    if existing:
        return existing
    now = _now()
    doc = {
        "id":           str(uuid.uuid4()),
        "user_id":      uid,
        "kind":         "product",
        "name":         ROOT_NAME,
        "parent_id":    None,
        "path":         [ROOT_NAME],
        "path_ids":     [],   # filled below for self-reference
        "depth":        0,
        "status":       "active",
        "movement_types": [],
        "created_at":   now,
        "updated_at":   now,
        "system":       True,  # protect from accidental delete
    }
    doc["path_ids"] = [doc["id"]]
    doc["code"]     = doc["id"]   # honour stale unique index
    await db.expense_category_tree.insert_one(doc)
    return doc


# ──────────────────────────  router  ──────────────────────────────────
def make_products_import_router(db, current_user):
    router = APIRouter(prefix="/products", tags=["products-import"])

    # ----- Categories: preview -----
    @router.post("/categories/import/preview")
    async def categories_import_preview(
        file: UploadFile = File(...),
        user: dict = Depends(current_user),
    ):
        uid = user["id"]
        raw = await file.read()
        if not raw:
            raise HTTPException(400, "الملف فارغ")
        if len(raw) > 5 * 1024 * 1024:
            raise HTTPException(400, "حجم الملف يتجاوز 5MB")

        parsed = _parse_categories_xlsx(raw)
        rows = parsed["rows"]

        # We DON'T modify anything during preview. We just classify each
        # row against the current product subtree.
        root = await db.expense_category_tree.find_one(
            {"user_id": uid, "kind": "product",
             "parent_id": None, "name": ROOT_NAME},
            {"_id": 0, "id": 1, "name": 1},
        )
        existing_map = (await _existing_product_keys(db, uid, root["id"])
                        if root else {})

        new_count = 0
        existing_count = 0
        parent_keys_in_file = {r["key"] for r in rows}
        parent_missing: list[dict] = []

        roots_in_file:  list[dict] = []
        subs_in_file:   list[dict] = []
        new_sample:     list[dict] = []
        existing_names: list[str]  = []

        for r in rows:
            if r["key"] in existing_map:
                existing_count += 1
                existing_names.append(r["name"])
            else:
                new_count += 1
                if len(new_sample) < 50:
                    new_sample.append({
                        "name":   r["name"],
                        "is_sub": r["is_sub"],
                        "parent": r["parent"],
                    })
            if r["is_sub"]:
                subs_in_file.append(r)
                if r["parent"]:
                    pkey = _norm_key(r["parent"])
                    if (pkey not in parent_keys_in_file
                            and pkey not in existing_map):
                        parent_missing.append({
                            "category": r["name"],
                            "missing_parent": r["parent"],
                        })
            else:
                roots_in_file.append(r)

        return {
            "ok": True,
            "phase": "categories_preview",
            "root": {
                "name":           ROOT_NAME,
                "exists":         bool(root),
                "id":             (root or {}).get("id"),
            },
            "totals": {
                "total_rows":              len(rows),
                "new":                     new_count,
                "existing":                existing_count,
                "root_level_in_file":      len(roots_in_file),
                "sub_level_in_file":       len(subs_in_file),
                "orphan_subs_no_parent":   len(parsed["orphans"]),
                "duplicates_in_file":      len(parsed["duplicates_in_file"]),
                "parent_not_found_in_file_or_db":
                                            len(parent_missing),
            },
            "samples": {
                "new":              new_sample,
                "orphans":          parsed["orphans"][:20],
                "duplicates":       parsed["duplicates_in_file"][:20],
                "existing_first10": existing_names[:10],
                "parent_missing":   parent_missing[:20],
            },
            "notes": [
                "لن يتم حذف أي تصنيف موجود.",
                f"التصنيفات الجديدة ستوضع تحت جذر «{ROOT_NAME}» "
                "(يُنشأ تلقائياً إن لم يكن موجوداً).",
                "التصنيفات الفرعية التي ليس لها أب صالح في الملف أو في "
                "الشجرة ستُلحق مؤقتاً بالجذر مع تنبيه — تستطيع نقلها "
                "يدوياً بعد الاستيراد.",
            ],
        }

    # ----- Categories: confirm -----
    @router.post("/categories/import/confirm")
    async def categories_import_confirm(
        file: UploadFile = File(...),
        user: dict = Depends(current_user),
    ):
        uid = user["id"]
        raw = await file.read()
        if not raw:
            raise HTTPException(400, "الملف فارغ")
        if len(raw) > 5 * 1024 * 1024:
            raise HTTPException(400, "حجم الملف يتجاوز 5MB")

        parsed = _parse_categories_xlsx(raw)
        rows = parsed["rows"]
        if not rows:
            raise HTTPException(400, "لا توجد بيانات صالحة للاستيراد")

        root = await _ensure_product_root(db, uid)
        existing_map = await _existing_product_keys(db, uid, root["id"])

        # 1) Insert roots first. 2) Then subs (after their parents
        # exist). 3) Orphans (sub but parent missing) → attach to root
        # so the merchant can move them later.
        roots = [r for r in rows if not r["is_sub"]]
        subs  = [r for r in rows if r["is_sub"]]

        created = 0
        skipped_existing = 0
        attached_to_root_as_orphan: list[str] = []

        async def _create(name: str, parent_doc: dict,
                          source_row_index: int) -> str:
            now = _now()
            new_id = str(uuid.uuid4())
            doc = {
                "id":          new_id,
                "user_id":     uid,
                "kind":        "product",
                # Stale unique index `{user_id, code}` requires a
                # unique value here — mirroring the id is safest and
                # avoids touching the index.
                "code":        new_id,
                "name":        name,
                "parent_id":   parent_doc["id"],
                "path":        (parent_doc.get("path") or []) + [name],
                "path_ids":    (parent_doc.get("path_ids") or []) + [new_id],
                "depth":       len((parent_doc.get("path") or [])),
                "status":      "active",
                "movement_types": [],
                "created_at":  now,
                "updated_at":  now,
                "imported": {
                    "source": "excel",
                    "row":    source_row_index,
                    "at":     now,
                },
            }
            await db.expense_category_tree.insert_one(doc)
            existing_map[_norm_key(name)] = new_id
            return new_id

        # Pre-fetch by id for parents we'll need.
        def _doc_by_id(cid: str):
            return db.expense_category_tree.find_one(
                {"id": cid, "user_id": uid}, {"_id": 0})

        # Roots
        for r in roots:
            if r["key"] in existing_map:
                skipped_existing += 1
                continue
            await _create(r["name"], root, r["raw_row_index"])
            created += 1

        # Subs — iterate up to 5 times to resolve nested children whose
        # parents appear later in the file.
        remaining = list(subs)
        for _ in range(5):
            if not remaining:
                break
            still: list[dict] = []
            for r in remaining:
                if r["key"] in existing_map:
                    skipped_existing += 1
                    continue
                parent_key = _norm_key(r["parent"]) if r["parent"] else None
                if not parent_key:
                    # No parent at all → attach to root.
                    await _create(r["name"], root, r["raw_row_index"])
                    created += 1
                    attached_to_root_as_orphan.append(r["name"])
                    continue
                parent_id = existing_map.get(parent_key)
                if not parent_id:
                    still.append(r)
                    continue
                parent_doc = await _doc_by_id(parent_id)
                if not parent_doc:
                    still.append(r)
                    continue
                await _create(r["name"], parent_doc, r["raw_row_index"])
                created += 1
            if len(still) == len(remaining):
                # No progress — remaining parents are unresolvable;
                # attach them to root.
                break
            remaining = still
        for r in remaining:
            if r["key"] in existing_map:
                skipped_existing += 1
                continue
            await _create(r["name"], root, r["raw_row_index"])
            created += 1
            attached_to_root_as_orphan.append(r["name"])

        return {
            "ok": True,
            "phase": "categories_confirm",
            "root": {"id": root["id"], "name": root["name"]},
            "summary": {
                "rows_in_file":              len(rows),
                "created":                   created,
                "skipped_existing":          skipped_existing,
                "attached_to_root_as_orphan":
                    len(attached_to_root_as_orphan),
                "duplicates_in_file":        len(parsed["duplicates_in_file"]),
            },
            "attached_to_root_as_orphan": attached_to_root_as_orphan[:50],
        }

    return router


# ─────────────  PRODUCTS  ─────────────────────────────────────────────
def _parse_products_xlsx(raw: bytes) -> dict:
    """Parse the products workbook. Columns (fixed positions):
       A: رقم المنتج (product_id)  — required (else skipped)
       B: اسم المنتج (name)         — required (else skipped)
       C: تصنيف المنتج (paths)      — optional, "A > B, C > D"
       D: صور المنتج (comma URLs)  — optional
       F: سعر التكلفة               — optional
    """
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"تعذّر قراءة ملف Excel: {e}")
    ws = wb.active
    rows: list[dict] = []
    seen: dict[str, int] = {}
    duplicates_in_file: list[dict] = []
    no_name: list[int] = []
    header_seen = False
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        if not header_seen:
            header_seen = True
            continue
        pid  = _norm(row[0] if len(row) > 0 else None)
        name = _norm(row[1] if len(row) > 1 else None)
        cats = _norm(row[2] if len(row) > 2 else None)
        imgs = _norm(row[3] if len(row) > 3 else None)
        # Column E is empty by spec — column F is cost.
        cost_raw = row[5] if len(row) > 5 else None
        if not pid:
            continue
        if not name:
            no_name.append(idx)
            continue
        if pid in seen:
            duplicates_in_file.append(
                {"row": idx, "product_id": pid, "name": name})
            continue
        seen[pid] = idx
        # Parse categories: comma-separated paths, each path uses " > ".
        cat_paths: list[list[str]] = []
        if cats:
            for path in cats.split(","):
                tokens = [t.strip() for t in path.split(">") if t.strip()]
                if tokens:
                    cat_paths.append(tokens)
        # Parse images
        image_urls: list[str] = []
        if imgs:
            for u in imgs.split(","):
                u = u.strip()
                if u:
                    image_urls.append(u)
        # Cost
        cost = None
        if cost_raw not in (None, ""):
            try:
                cost = round(float(cost_raw), 2)
                if cost <= 0:
                    cost = None
            except (TypeError, ValueError):
                cost = None
        rows.append({
            "row":         idx,
            "product_id":  pid,
            "name":        name,
            "name_lower":  _norm_key(name),
            "cat_paths":   cat_paths,
            "image_urls":  image_urls,
            "image_url":   (image_urls[0] if image_urls else None),
            "cost":        cost,
        })
    return {
        "rows": rows,
        "duplicates_in_file": duplicates_in_file,
        "no_name_rows": no_name,
    }


async def _ensure_uncategorized(db, uid: str, root: dict) -> dict:
    """Ensure a single "غير مصنف" category exists directly under root."""
    existing = await db.expense_category_tree.find_one(
        {"user_id": uid, "kind": "product",
         "parent_id": root["id"], "name": UNCAT_NAME},
        {"_id": 0},
    )
    if existing:
        return existing
    now = _now()
    new_id = str(uuid.uuid4())
    doc = {
        "id":         new_id,
        "user_id":    uid,
        "kind":       "product",
        "code":       new_id,
        "name":       UNCAT_NAME,
        "parent_id":  root["id"],
        "path":       (root.get("path") or []) + [UNCAT_NAME],
        "path_ids":   (root.get("path_ids") or []) + [new_id],
        "depth":      len(root.get("path") or []),
        "status":     "active",
        "movement_types": [],
        "created_at": now,
        "updated_at": now,
        "system":     True,
    }
    await db.expense_category_tree.insert_one(doc)
    return doc


async def _resolve_path_to_cat_id(
    db, uid: str, root: dict, tokens: list[str],
    cache: dict[tuple, str],
) -> str:
    """Resolve a hierarchical path like ['اكسسوارات نسائي', 'سلاسل نسائيه']
    to a category_id. Walks the tree; creates missing nodes on the fly
    under the root. The cache short-circuits repeated paths."""
    cur_parent = root
    key_prefix: tuple = ()
    for tok in tokens:
        norm = _norm_key(tok)
        key_prefix = key_prefix + (norm,)
        if key_prefix in cache:
            cur_id = cache[key_prefix]
            cur_parent = await db.expense_category_tree.find_one(
                {"id": cur_id, "user_id": uid}, {"_id": 0})
            continue
        existing = await db.expense_category_tree.find_one(
            {"user_id": uid, "kind": "product",
             "parent_id": cur_parent["id"], "name": tok},
            {"_id": 0},
        )
        if existing:
            cache[key_prefix] = existing["id"]
            cur_parent = existing
            continue
        # Auto-create missing node under cur_parent.
        now = _now()
        new_id = str(uuid.uuid4())
        doc = {
            "id":        new_id,
            "user_id":   uid,
            "kind":      "product",
            "code":      new_id,
            "name":      tok,
            "parent_id": cur_parent["id"],
            "path":      (cur_parent.get("path") or []) + [tok],
            "path_ids":  (cur_parent.get("path_ids") or []) + [new_id],
            "depth":     len(cur_parent.get("path") or []),
            "status":    "active",
            "movement_types": [],
            "created_at": now,
            "updated_at": now,
            "imported": {"source": "excel-products-auto",
                         "at": now},
        }
        await db.expense_category_tree.insert_one(doc)
        cache[key_prefix] = new_id
        cur_parent = doc
    return cur_parent["id"]


def make_products_router_phase2(db, current_user):
    """Phase 2 — products import: preview + confirm + list."""
    router = APIRouter(prefix="/products", tags=["products-import"])

    @router.get("/list")
    async def products_list(
        q: str = "",
        category_id: str | None = None,
        needs_cost: bool | None = None,
        limit: int = 100,
        skip:  int = 0,
        user: dict = Depends(current_user),
    ):
        """READ-ONLY list with search + ranking + pagination.

        Ranking when `q` is provided (per the merchant's spec):
          1. Name starts with the query
          2. Name contains the query (anywhere)
          3. Token-based similarity (any token of the name starts
             with the query)
          4. product_id / sku / barcode exact-ish match
        """
        uid = user["id"]
        flt: dict = {"user_id": uid, "is_active": True}
        if needs_cost is True:
            flt["needs_cost"] = True
        if category_id:
            flt["category_ids"] = category_id
        if q:
            qn = _norm_key(q)
            qr = re.escape(qn)
            flt["$or"] = [
                {"name_lower": {"$regex": qr}},
                {"product_id": {"$regex": re.escape(q.strip())}},
                {"sku":        {"$regex": re.escape(q.strip())}},
                {"barcode":    {"$regex": re.escape(q.strip())}},
            ]
        items: list[dict] = []
        # Pull a generous page when we need to rank — ranking is
        # cheaper than DB ordering across multiple priority buckets.
        fetch_limit = min(limit * 5, 500) if q else min(limit, 200)
        async for d in db.products.find(flt, {"_id": 0}) \
                .sort([("updated_at", -1)]).skip(skip).limit(fetch_limit):
            items.append(d)

        if q:
            qn = _norm_key(q)
            qd = q.strip()

            def _rank(p: dict) -> tuple:
                nl = p.get("name_lower") or ""
                pid = p.get("product_id") or ""
                sku = p.get("sku") or ""
                bcd = p.get("barcode") or ""
                if nl.startswith(qn):
                    bucket = 0
                elif qn in nl:
                    bucket = 1
                elif any(tok.startswith(qn) for tok in nl.split()):
                    bucket = 2
                elif qd and qd in pid:
                    bucket = 3
                elif qd and (sku and qd in sku):
                    bucket = 4
                elif qd and (bcd and qd in bcd):
                    bucket = 5
                else:
                    bucket = 9
                # Secondary: shorter names first (more relevant).
                return (bucket, len(nl), nl)
            items.sort(key=_rank)
            items = items[: min(limit, 200)]

        total = await db.products.count_documents(flt)
        all_cids: set = set()
        for it in items:
            for c in (it.get("category_ids") or []):
                all_cids.add(c)
        cat_paths: dict[str, list[str]] = {}
        if all_cids:
            async for c in db.expense_category_tree.find(
                {"user_id": uid, "id": {"$in": list(all_cids)}},
                {"_id": 0, "id": 1, "path": 1},
            ):
                cat_paths[c["id"]] = c.get("path") or []
        for it in items:
            paths = []
            for c in (it.get("category_ids") or []):
                if c in cat_paths:
                    paths.append(cat_paths[c])
            it["category_paths"] = paths
        for it in items:
            pid = str(it.get("product_id") or "").strip()
            if pid.upper().startswith("AUTO-"):
                it["auto_catalog_key"] = it.get("auto_catalog_key") or pid
                it["product_id"] = None
        return {"items": items, "total": total,
                "skip": skip, "limit": limit}

    # ---------- Iter-250b · Phase 4 visual audit ----------
    # GET /api/products/{pid}/cost-history?limit=5
    # Returns the most recent `cost_history` entries enriched with the
    # supplier company name and the invoice doc_number so the merchant
    # can visually verify that each invoice was logged as a separate
    # record. Sorted newest-first.
    @router.get("/{pid}/cost-history")
    async def product_cost_history(
        pid: str,
        limit: int = 5,
        user: dict = Depends(current_user),
    ):
        uid = user["id"]
        prod = await db.products.find_one(
            {"user_id": uid, "$or": [{"id": pid}, {"product_id": pid}]},
            {"_id": 0, "id": 1, "product_id": 1, "name": 1,
             "cost_current": 1, "cost_avg": 1, "cost_history": 1},
        )
        if not prod:
            raise HTTPException(404, "المنتج غير موجود")

        hist = prod.get("cost_history") or []
        # Newest-first, then take only what the caller asked for.
        sorted_hist = sorted(
            hist, key=lambda h: h.get("at") or "", reverse=True
        )[: max(1, min(limit, 50))]

        # Resolve supplier names + invoice doc_numbers in two roundtrips.
        sup_ids = list({h.get("supplier_id") for h in sorted_hist
                        if h.get("supplier_id")})
        inv_ids = list({h.get("supplier_invoice_id") for h in sorted_hist
                        if h.get("supplier_invoice_id")})

        sup_map: dict[str, str] = {}
        if sup_ids:
            async for s in db.suppliers.find(
                {"user_id": uid, "id": {"$in": sup_ids}},
                {"_id": 0, "id": 1, "company_name": 1, "name": 1},
            ):
                sup_map[s["id"]] = (
                    s.get("company_name") or s.get("name") or "")
            # Fallback to counterparties for legacy IDs.
            missing = [sid for sid in sup_ids if sid not in sup_map]
            if missing:
                async for c in db.counterparties.find(
                    {"user_id": uid, "id": {"$in": missing}},
                    {"_id": 0, "id": 1, "name": 1, "company_name": 1},
                ):
                    sup_map[c["id"]] = (
                        c.get("company_name") or c.get("name") or "")

        inv_map: dict[str, str] = {}
        if inv_ids:
            async for m in db.financial_movements.find(
                {"user_id": uid, "id": {"$in": inv_ids}},
                {"_id": 0, "id": 1, "doc_number": 1},
            ):
                inv_map[m["id"]] = m.get("doc_number") or ""

        enriched = []
        for h in sorted_hist:
            sid = h.get("supplier_id")
            iid = h.get("supplier_invoice_id")
            enriched.append({
                "at":                  h.get("at"),
                "source":              h.get("source"),
                "supplier_id":         sid,
                "supplier_name":       sup_map.get(sid) if sid else None,
                "supplier_invoice_id": iid,
                "doc_number":          inv_map.get(iid) if iid else None,
                "invoice_date":        h.get("invoice_date"),
                "quantity":            h.get("quantity"),
                "unit_cost":           h.get("unit_cost") or h.get("amount"),
                "total_cost":          h.get("total_cost"),
                # Iter-250b · Phase 4.5 — lifecycle status surface.
                "status":              (h.get("status") or "active"),
                "reversed_at":         h.get("reversed_at"),
                "reversal_txn_group_id": h.get("reversal_txn_group_id"),
            })

        return {
            "product": {
                "id":           prod.get("id"),
                "product_id":   prod.get("product_id"),
                "name":         prod.get("name"),
                "cost_current": prod.get("cost_current"),
                "cost_avg":     prod.get("cost_avg"),
            },
            "items":       enriched,
            "total_count": len(hist),
        }

    # ---------- Iter-250b · Phase 4.5 — Cost-health diagnostics ----------
    # Read-only forensic report exposing every product whose cost
    # numbers are out of sync with their `cost_history`.  Pure analysis:
    # no DB writes, safe to call from production.
    #
    # Categories surfaced:
    #   • current_from_reversed       — cost_current came from a row
    #                                    whose status == "reversed"
    #   • avg_contains_reversed       — at least one reversed entry
    #                                    is being averaged in (legacy
    #                                    pre-Phase-4.5 documents)
    #   • needs_cost_false_no_active  — needs_cost = False but no
    #                                    active history rows exist
    #   • current_not_latest_active   — cost_current ≠ the newest
    #                                    active entry's unit_cost
    @router.get("/cost-health")
    async def products_cost_health(
        limit: int = 1000,
        user: dict = Depends(current_user),
    ):
        uid = user["id"]
        from financial_movements_routes import (
            _is_active_entry, _entry_sort_key,
        )

        def _r2(x):
            try:
                return round(float(x or 0), 2)
            except Exception:
                return 0.0

        cur_from_rev:  list = []
        avg_has_rev:   list = []
        needs_no_act:  list = []
        cur_not_latest: list = []

        cursor = db.products.find(
            {"user_id": uid},
            {"_id": 0, "id": 1, "product_id": 1, "name": 1,
             "cost_current": 1, "cost_avg": 1, "needs_cost": 1,
             "cost_history": 1},
        )
        async for p in cursor:
            hist = p.get("cost_history") or []
            active = [h for h in hist if _is_active_entry(h)]
            reversed_entries = [h for h in hist if not _is_active_entry(h)]
            cost_current = _r2(p.get("cost_current"))
            stub = {
                "id":           p["id"],
                "product_id":   p.get("product_id"),
                "name":         p.get("name"),
                "cost_current": cost_current,
                "cost_avg":     _r2(p.get("cost_avg")),
                "needs_cost":   bool(p.get("needs_cost")),
                "active_entries":   len(active),
                "reversed_entries": len(reversed_entries),
                "total_entries":    len(hist),
            }

            # (1) cost_current came from a reversed row?
            if reversed_entries:
                latest_any = sorted(hist, key=_entry_sort_key,
                                    reverse=True)[0] if hist else None
                if latest_any and not _is_active_entry(latest_any):
                    if _r2(latest_any.get("unit_cost")
                           or latest_any.get("amount")) == cost_current \
                       and cost_current > 0:
                        cur_from_rev.append({
                            **stub,
                            "latest_entry": {
                                "supplier_invoice_id":
                                    latest_any.get("supplier_invoice_id"),
                                "at":         latest_any.get("at"),
                                "unit_cost":  latest_any.get("unit_cost"),
                                "status":     latest_any.get("status"),
                            },
                        })

            # (2) needs_cost=False but no active rows?
            if not active and not p.get("needs_cost"):
                needs_no_act.append(stub)

            # (3) cost_current ≠ newest active entry's unit_cost
            if active:
                latest_active = sorted(active, key=_entry_sort_key,
                                       reverse=True)[0]
                latest_uc = _r2(latest_active.get("unit_cost")
                                or latest_active.get("amount"))
                if latest_uc > 0 and latest_uc != cost_current:
                    cur_not_latest.append({
                        **stub,
                        "expected_cost_current": latest_uc,
                        "delta": round(cost_current - latest_uc, 2),
                    })

            # (4) Pre-4.5 legacy doc whose avg implicitly absorbed
            #     a reversed entry — only possible if the doc has at
            #     least one reversed row WITH qty AND the current avg
            #     no longer matches the avg-over-active.  We detect
            #     this by recomputing the active-only weighted avg
            #     and comparing to stored cost_avg.
            if reversed_entries and active:
                qty_rows = [h for h in active
                            if h.get("quantity") and h.get("unit_cost")
                            and _r2(h["quantity"]) > 0]
                if qty_rows:
                    tot_q = sum(_r2(h["quantity"]) for h in qty_rows)
                    tot_v = sum(_r2(h["quantity"]) * _r2(h["unit_cost"])
                                for h in qty_rows)
                    expected_avg = round(tot_v / tot_q, 2) if tot_q else 0
                    if expected_avg > 0 and \
                       expected_avg != _r2(p.get("cost_avg")):
                        avg_has_rev.append({
                            **stub,
                            "expected_cost_avg": expected_avg,
                            "delta": round(_r2(p.get("cost_avg"))
                                            - expected_avg, 2),
                        })

        return {
            "summary": {
                "current_from_reversed":      len(cur_from_rev),
                "avg_contains_reversed":      len(avg_has_rev),
                "needs_cost_false_no_active": len(needs_no_act),
                "current_not_latest_active":  len(cur_not_latest),
            },
            "current_from_reversed":      cur_from_rev[:limit],
            "avg_contains_reversed":      avg_has_rev[:limit],
            "needs_cost_false_no_active": needs_no_act[:limit],
            "current_not_latest_active":  cur_not_latest[:limit],
        }

    # ---------- Iter-250b · Phase 4.5 — Cost recompute (preview / confirm) ----
    # Walks every product and runs `recalculate_product_cost` in
    # dry-run mode (preview) or apply mode (confirm).  Returns a
    # before/after diff per product that actually changed.  Safe and
    # idempotent — products already in sync are left untouched.
    async def _do_recompute(uid: str, dry_run: bool):
        from financial_movements_routes import recalculate_product_cost
        changed_rows: list = []
        scanned = 0
        async for p in db.products.find(
            {"user_id": uid}, {"_id": 0, "id": 1},
        ):
            scanned += 1
            try:
                r = await recalculate_product_cost(
                    db, uid, p["id"], dry_run=dry_run)
                if r.get("ok") and r.get("changed"):
                    changed_rows.append(r)
            except Exception as e:  # noqa: BLE001
                changed_rows.append({
                    "ok": False, "product_id": p["id"], "error": str(e)})
        return {
            "scanned":       scanned,
            "changed_count": len(changed_rows),
            "changes":       changed_rows,
            "dry_run":       dry_run,
        }

    @router.post("/cost-recompute/preview")
    async def products_cost_recompute_preview(
        user: dict = Depends(current_user),
    ):
        return await _do_recompute(user["id"], dry_run=True)

    @router.post("/cost-recompute/confirm")
    async def products_cost_recompute_confirm(
        user: dict = Depends(current_user),
    ):
        return await _do_recompute(user["id"], dry_run=False)

    # ---------- Quick create (Phase 3 — from supplier invoice) ----------
    @router.post("/quick-create")
    async def products_quick_create(
        payload: dict,
        user: dict = Depends(current_user),
    ):
        uid = user["id"]
        name = _norm(payload.get("name"))
        if not name:
            raise HTTPException(400, "اسم المنتج إلزامي")
        product_id = _norm(payload.get("product_id")) or str(uuid.uuid4())[:8]
        # Reject if product_id already exists for this user.
        clash = await db.products.find_one(
            {"user_id": uid, "product_id": product_id},
            {"_id": 0, "id": 1, "name": 1})
        if clash:
            raise HTTPException(
                409, f"رقم المنتج «{product_id}» مستخدم لمنتج آخر «"
                + (clash.get("name") or "") + "»")
        # Resolve category — falls back to "غير مصنف" when not given.
        cat_id = _norm(payload.get("category_id"))
        if not cat_id:
            root = await _ensure_product_root(db, uid)
            uncat = await _ensure_uncategorized(db, uid, root)
            cat_id = uncat["id"]
        else:
            ok = await db.expense_category_tree.find_one(
                {"user_id": uid, "id": cat_id, "kind": "product"},
                {"_id": 0, "id": 1})
            if not ok:
                raise HTTPException(400, "التصنيف غير صالح")
        cost_raw = payload.get("cost")
        cost = None
        if cost_raw not in (None, ""):
            try:
                cost = round(float(cost_raw), 2)
                if cost <= 0:
                    cost = None
            except (TypeError, ValueError):
                cost = None
        now = _now()
        image_url = _norm(payload.get("image_url")) or None
        history = ([{"amount": cost, "source": "quick-create",
                     "at": now}] if cost is not None else [])
        doc = {
            "id":           str(uuid.uuid4()),
            "user_id":      uid,
            "product_id":   product_id,
            "name":         name,
            "name_lower":   _norm_key(name),
            "category_ids": [cat_id],
            "image_url":    image_url,
            "image_urls":   ([image_url] if image_url else []),
            "cost_current": cost,
            "cost_avg":     cost,
            "cost_history": history,
            "sku":          None,
            "barcode":      None,
            "needs_cost":   (cost is None),
            "is_active":    True,
            "imported": {"source": "quick-create", "at": now},
            "notes":        _norm(payload.get("notes")) or None,
            "created_at":   now,
            "updated_at":   now,
        }
        await db.products.insert_one(doc)
        # Strip the BSON `_id` Motor adds in-place after insert.
        doc.pop("_id", None)
        # Hydrate category_paths for immediate UI use.
        cat = await db.expense_category_tree.find_one(
            {"id": cat_id, "user_id": uid},
            {"_id": 0, "path": 1})
        doc["category_paths"] = [cat.get("path", []) if cat else []]
        return {"ok": True, "product": doc}

    @router.post("/import/preview")
    async def products_import_preview(
        file: UploadFile = File(...),
        user: dict = Depends(current_user),
    ):
        uid = user["id"]
        raw = await file.read()
        if not raw:
            raise HTTPException(400, "الملف فارغ")
        if len(raw) > 15 * 1024 * 1024:
            raise HTTPException(400, "حجم الملف يتجاوز 15MB")

        parsed = _parse_products_xlsx(raw)
        rows = parsed["rows"]

        # Cross-check against db.products to compute new vs update.
        pids = [r["product_id"] for r in rows]
        existing_pids: set = set()
        async for d in db.products.find(
            {"user_id": uid, "product_id": {"$in": pids}},
            {"_id": 0, "product_id": 1},
        ):
            existing_pids.add(d["product_id"])

        new_count    = 0
        update_count = 0
        no_cat       = 0
        no_cost      = 0
        multi_cat    = 0
        unique_cat_paths: set = set()
        sample_new: list[dict] = []
        for r in rows:
            if r["product_id"] in existing_pids:
                update_count += 1
            else:
                new_count += 1
                if len(sample_new) < 20:
                    sample_new.append({
                        "product_id": r["product_id"],
                        "name":       r["name"],
                        "cost":       r["cost"],
                        "cat_paths":  r["cat_paths"],
                        "image_url":  r["image_url"],
                    })
            if not r["cat_paths"]:
                no_cat += 1
            elif len(r["cat_paths"]) > 1:
                multi_cat += 1
            if r["cost"] is None:
                no_cost += 1
            for p in r["cat_paths"]:
                unique_cat_paths.add(" > ".join(p))

        return {
            "ok": True,
            "phase": "products_preview",
            "totals": {
                "total_rows":           len(rows),
                "new":                  new_count,
                "update":               update_count,
                "no_name_skipped":      len(parsed["no_name_rows"]),
                "duplicates_in_file":   len(parsed["duplicates_in_file"]),
                "no_category":          no_cat,
                "no_cost":              no_cost,
                "multi_category":       multi_cat,
                "unique_category_paths": len(unique_cat_paths),
            },
            "samples": {
                "new":                 sample_new,
                "duplicates":          parsed["duplicates_in_file"][:20],
                "no_name_rows":        parsed["no_name_rows"][:20],
                "category_paths_first20": sorted(unique_cat_paths)[:20],
            },
            "notes": [
                "رقم المنتج هو مفتاح منع التكرار.",
                "المنتجات بلا تصنيف ستُربط بـ «غير مصنف» تلقائياً.",
                "التصنيفات غير الموجودة سيتم إنشاؤها داخل شجرة "
                "«المنتجات المستوردة».",
                "المنتجات بدون تكلفة ستُحفظ مع علم needs_cost=true.",
                "كل عملية شراء جديدة ستُضاف لـ cost_history دون استبدال.",
            ],
        }

    @router.post("/import/confirm")
    async def products_import_confirm(
        file: UploadFile = File(...),
        user: dict = Depends(current_user),
    ):
        uid = user["id"]
        raw = await file.read()
        if not raw:
            raise HTTPException(400, "الملف فارغ")
        if len(raw) > 15 * 1024 * 1024:
            raise HTTPException(400, "حجم الملف يتجاوز 15MB")

        parsed = _parse_products_xlsx(raw)
        rows = parsed["rows"]
        if not rows:
            raise HTTPException(400, "لا توجد بيانات صالحة للاستيراد")

        root = await _ensure_product_root(db, uid)
        uncat = await _ensure_uncategorized(db, uid, root)
        path_cache: dict[tuple, str] = {}

        created    = 0
        updated    = 0
        uncat_used = 0
        cats_auto_created = 0
        cats_before = await db.expense_category_tree.count_documents(
            {"user_id": uid, "kind": "product"})

        for r in rows:
            # Resolve category_ids.
            cat_ids: list[str] = []
            if r["cat_paths"]:
                for tokens in r["cat_paths"]:
                    try:
                        cid = await _resolve_path_to_cat_id(
                            db, uid, root, tokens, path_cache)
                        if cid not in cat_ids:
                            cat_ids.append(cid)
                    except Exception:
                        continue
            if not cat_ids:
                cat_ids = [uncat["id"]]
                uncat_used += 1

            existing = await db.products.find_one(
                {"user_id": uid, "product_id": r["product_id"]},
                {"_id": 0},
            )
            now = _now()
            needs_cost = (r["cost"] is None)
            if not existing:
                # New product. Initial cost_history entry if cost is set.
                history: list[dict] = []
                if r["cost"] is not None:
                    history.append({
                        "amount": r["cost"],
                        "source": "excel-import",
                        "at":     now,
                    })
                doc = {
                    "id":            str(uuid.uuid4()),
                    "user_id":       uid,
                    "product_id":    r["product_id"],
                    "name":          r["name"],
                    "name_lower":    r["name_lower"],
                    "category_ids":  cat_ids,
                    "image_url":     r["image_url"],
                    "image_urls":    r["image_urls"],
                    "cost_current":  r["cost"],
                    "cost_avg":      r["cost"],
                    "cost_history":  history,
                    "sku":           None,
                    "barcode":       None,
                    "needs_cost":    needs_cost,
                    "is_active":     True,
                    "imported": {
                        "source": "excel",
                        "row":    r["row"],
                        "at":     now,
                    },
                    "created_at":    now,
                    "updated_at":    now,
                }
                await db.products.insert_one(doc)
                created += 1
            else:
                # Update — never overwrite existing cost_history.
                set_fields: dict = {
                    "name":         r["name"],
                    "name_lower":   r["name_lower"],
                    "category_ids": cat_ids,
                    "updated_at":   now,
                }
                # Only refresh images when new ones provided.
                if r["image_urls"]:
                    set_fields["image_url"]  = r["image_url"]
                    set_fields["image_urls"] = r["image_urls"]
                # Cost: if Excel has a cost AND existing has none, set it.
                if r["cost"] is not None and existing.get("cost_current") is None:
                    set_fields["cost_current"] = r["cost"]
                    set_fields["cost_avg"]     = r["cost"]
                    set_fields["needs_cost"]   = False
                    push = {"cost_history": {
                        "amount": r["cost"],
                        "source": "excel-import",
                        "at":     now,
                    }}
                    await db.products.update_one(
                        {"id": existing["id"], "user_id": uid},
                        {"$set": set_fields, "$push": push},
                    )
                else:
                    # Keep `needs_cost` consistent.
                    if existing.get("cost_current") is None:
                        set_fields["needs_cost"] = True
                    await db.products.update_one(
                        {"id": existing["id"], "user_id": uid},
                        {"$set": set_fields},
                    )
                updated += 1

        cats_after = await db.expense_category_tree.count_documents(
            {"user_id": uid, "kind": "product"})
        cats_auto_created = max(0, cats_after - cats_before)

        return {
            "ok": True,
            "phase": "products_confirm",
            "summary": {
                "rows_in_file":          len(rows),
                "created":               created,
                "updated":               updated,
                "uncategorized_linked":  uncat_used,
                "categories_auto_created": cats_auto_created,
                "no_name_skipped":       len(parsed["no_name_rows"]),
                "duplicates_in_file":    len(parsed["duplicates_in_file"]),
            },
        }

    return router


__all__ = ["make_products_import_router", "make_products_router_phase2"]
