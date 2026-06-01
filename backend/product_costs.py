"""Product Cost Management — iteration 19.

Provides CRUD endpoints for a merchant-maintained catalogue of product
purchase costs (per SKU + supplier), plus integration helpers that the
Make.com webhook ingestion and dashboard summary call to compute the
TRUE net profit of each order.

Lookup strategy (per merchant requirement):
    1. Match by SKU (case-insensitive, primary key — most stable in Salla).
    2. Fall back to Salla product_id.
    3. If neither matches → the order line item is flagged in
       `missing_product_cost_lines` (list returned with each ingestion +
       persisted on the order doc so the dashboard can show
       "X منتجات بدون تكلفة").

Endpoints registered under `/api/product-costs`:
    GET    /                    paginated catalogue (search + filter)
    POST   /                    create one item
    PUT    /{id}                update fields
    DELETE /{id}                soft-delete (sets is_active=False)
    POST   /import              Excel upload (bulk insert/update)
    GET    /missing             order lines with no matching cost entry
    GET    /summary             today / month / avg / top profitable
    POST   /recompute           recompute persisted cost on every existing
                                order (used after import or bulk edits)

Collections:
    product_costs    — catalogue (one doc per (user_id, sku))
    unified_orders   — enriched per-line `cost_price` + per-order
                       `total_product_cost` + `missing_product_cost_lines`
                       (added in this module, never overwritten by other
                       writers).
"""
from __future__ import annotations

import io
import logging
import re
import uuid
from datetime import datetime, timezone, timedelta
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field, model_validator

logger = logging.getLogger(__name__)


# ── Helpers ────────────────────────────────────────────────────────────────
def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm_sku(value: Any) -> str:
    """Normalise a SKU: strip, upper-case so 'neck001' == 'NECK001'."""
    return str(value or "").strip().upper()


def _norm_product_id(value: Any) -> str:
    return str(value or "").strip()


def _to_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


# ── Public helper used by webhook ingestion + dashboard ────────────────────
async def compute_order_cost(
    db, user_id: str, products: list,
) -> tuple[float, list[dict], list[dict]]:
    """Look up cost for every product line in an order.

    Returns:
        total_cost   — sum of (cost_price * quantity) across MATCHED lines, in SAR.
                       Note: this is the PARTIAL sum — unmatched lines are
                       NOT counted (per merchant rule: never assume 0 cost
                       for a missing product, flag the order instead).
        items        — per-line breakdown:
                       [{sku, product_id, name, quantity, unit_cost,
                         line_cost, matched_by: 'sku'|'product_id'|None}]
        missing      — lines with no cost match:
                       [{sku, product_id, name, quantity, image_url}]
    """
    total = 0.0
    items: list[dict] = []
    missing: list[dict] = []
    if not products:
        return 0.0, items, missing

    # Bulk-fetch all costs the user might need (faster than N round-trips).
    skus = {_norm_sku(p.get("sku")) for p in products if p.get("sku")}
    pids = {_norm_product_id(p.get("product_id") or p.get("id"))
            for p in products if (p.get("product_id") or p.get("id"))}
    or_clauses: list[dict] = []
    if skus:
        or_clauses.append({"sku_normalized": {"$in": list(skus)}})
    if pids:
        or_clauses.append({"product_id": {"$in": list(pids)}})
    cost_docs: list[dict] = []
    if or_clauses:
        cost_docs = await db.product_costs.find(
            # Iteration 25: skip rows whose cost is still pending (set
            # only by import when cost_price was empty). They must NOT
            # be treated as a match — the order should land in the
            # missing-cost list until the merchant fills in the price.
            {"user_id": user_id, "is_active": True,
             "$or": or_clauses,
             "cost_pending": {"$ne": True}},
            {"_id": 0},
        ).to_list(500)
    by_sku = {_norm_sku(d.get("sku")): d for d in cost_docs if d.get("sku")}
    by_pid = {_norm_product_id(d.get("product_id")): d
              for d in cost_docs if d.get("product_id")}

    for p in products:
        qty = _to_float(p.get("quantity"), 1.0)
        sku = _norm_sku(p.get("sku"))
        pid = _norm_product_id(p.get("product_id") or p.get("id"))
        name = (p.get("name") or "").strip()
        # Iteration 24: also pull the product's image_url from the
        # webhook payload so the "missing products" UI can render a
        # thumbnail for each unmatched product.
        image_url = str(p.get("image_url") or p.get("image") or "").strip()
        cost_doc = None
        matched_by = None
        if sku and sku in by_sku:
            cost_doc = by_sku[sku]
            matched_by = "sku"
        elif pid and pid in by_pid:
            cost_doc = by_pid[pid]
            matched_by = "product_id"
        if cost_doc:
            unit_cost = round(_to_float(cost_doc.get("cost_price")), 2)
            line_cost = round(unit_cost * qty, 2)
            total += line_cost
            items.append({
                "sku": sku or "", "product_id": pid or "",
                "name": name, "quantity": qty,
                "unit_cost": unit_cost, "line_cost": line_cost,
                "currency": cost_doc.get("currency") or "SAR",
                "supplier_name": cost_doc.get("supplier_name") or "",
                "matched_by": matched_by,
            })
        else:
            items.append({
                "sku": sku or "", "product_id": pid or "",
                "name": name, "quantity": qty,
                "unit_cost": 0.0, "line_cost": 0.0,
                "matched_by": None,
            })
            missing.append({
                "sku": sku or "", "product_id": pid or "",
                "name": name, "quantity": qty,
                "image_url": image_url,
            })
    return round(total, 2), items, missing


def _classify_profit_status(products: list, missing: list) -> str:
    """Iteration 24: classify an order's profit completeness.

    Returns one of:
      - "complete"               : has products[] AND all matched cost.
      - "incomplete_missing_cost": has products[] but ≥1 line has no cost.
      - "incomplete_no_products" : products[] is empty (typical Excel order
                                   or Make.com payload missing the array).
    """
    if not products:
        return "incomplete_no_products"
    if missing:
        return "incomplete_missing_cost"
    return "complete"


async def attach_cost_to_order_doc(db, user_id: str, order_doc: dict) -> dict:
    """Enrich an order doc with `cost_items`, `total_product_cost`,
    `missing_product_cost_lines`, and `profit_status` (iteration 24 —
    idempotent — safe to call on every upsert). Returns a $set patch
    dict so the caller can merge it into its update query.

    `total_product_cost` is the PARTIAL sum of matched lines only — we
    never assume 0 for a missing product. The merchant adds the cost
    via the "Missing Products" UI, then `attach_cost_to_order_doc` is
    re-run on all affected orders (see `_reprocess_orders_for_keys`).
    """
    products = order_doc.get("products") or []
    total, items, missing = await compute_order_cost(db, user_id, products)
    status = _classify_profit_status(products, missing)
    return {
        "total_product_cost": total,
        "cost_items": items,
        "missing_product_cost_lines": missing,
        "profit_status": status,
        "products_total_lines": len(products),
        "products_matched_lines": len(products) - len(missing),
        "cost_computed_at": _now_iso(),
    }


async def _reprocess_orders_for_keys(
    db, user_id: str, skus: set, pids: set,
) -> int:
    """Iteration 24: targeted re-run of `attach_cost_to_order_doc` for
    every order that currently has at least one missing line matching
    one of the supplied SKUs or product_ids. Used after the merchant
    creates/updates a `product_costs` entry from the "Missing Products"
    UI so prior orders flip from `incomplete_missing_cost` → `complete`
    without waiting for a full bulk recompute.

    Returns the number of orders re-enriched.
    """
    if not skus and not pids:
        return 0
    or_clauses: list[dict] = []
    if skus:
        # Both `missing_product_cost_lines.sku` AND `cost_items.sku` are
        # stored already-normalised (upper-case) by compute_order_cost.
        # We query BOTH because:
        #   - "missing" matches orders currently unmatched (covers POST
        #     /product-costs/ — merchant just added the cost).
        #   - "cost_items" matches orders ALREADY matched (covers PUT
        #     /product-costs/{id} — merchant edited the price).
        sku_list = [s for s in skus if s]
        if sku_list:
            or_clauses.append(
                {"missing_product_cost_lines.sku": {"$in": sku_list}}
            )
            or_clauses.append(
                {"cost_items.sku": {"$in": sku_list}}
            )
    if pids:
        pid_list = [p for p in pids if p]
        if pid_list:
            or_clauses.append(
                {"missing_product_cost_lines.product_id": {"$in": pid_list}}
            )
            or_clauses.append(
                {"cost_items.product_id": {"$in": pid_list}}
            )
    if not or_clauses:
        return 0
    affected = 0
    async for o in db.unified_orders.find(
        {"user_id": user_id, "$or": or_clauses},
        {"_id": 0, "order_number": 1, "products": 1},
    ):
        patch = await attach_cost_to_order_doc(db, user_id, o)
        await db.unified_orders.update_one(
            {"user_id": user_id, "order_number": o["order_number"]},
            {"$set": patch},
        )
        affected += 1
    return affected


async def _recompute_recent_orders(db, user_id: str, days: int = 2) -> int:
    """Iteration 26: unconditional recompute of EVERY order in the last
    `days` days, regardless of SKU. Called after every cost mutation
    (create/update/bulk-import) so the dashboard + reports immediately
    reflect updated costs without waiting for a manual `/recompute`.

    Idempotent — safe to call repeatedly. Complements the targeted
    `_reprocess_orders_for_keys` pass by also catching:
      • orders whose cost_items array was somehow inconsistent
      • freshly-ingested orders that arrived between two cost edits
    """
    if days <= 0:
        return 0
    cutoff = (datetime.now(timezone.utc).date()
              - timedelta(days=days - 1)).isoformat()
    updated = 0
    async for o in db.unified_orders.find(
        {"user_id": user_id, "order_date": {"$gte": cutoff}},
        {"_id": 0, "order_number": 1, "products": 1},
    ):
        patch = await attach_cost_to_order_doc(db, user_id, o)
        await db.unified_orders.update_one(
            {"user_id": user_id, "order_number": o["order_number"]},
            {"$set": patch},
        )
        updated += 1
    return updated


# ── Pydantic models ────────────────────────────────────────────────────────
class ProductCostIn(BaseModel):
    # Iteration 25: SKU is now OPTIONAL (Salla product Excel exports
    # frequently omit it). product_id is the primary identifier when
    # SKU is absent. At least one of {sku, product_id} MUST be present.
    sku: Optional[str] = Field(default="", max_length=80)
    product_id: Optional[str] = Field(default="", max_length=80)
    product_name: str = Field(min_length=1, max_length=200)
    supplier_name: Optional[str] = ""
    supplier_country: Optional[str] = ""
    supplier_notes: Optional[str] = ""
    # Iteration 25: cost_price is also optional now. When empty, the row
    # is saved as "pending cost" (cost_pending=True) and appears in the
    # `/missing` list — never assumed = 0 for profit calculations.
    cost_price: Optional[float] = Field(default=None, ge=0)
    currency: str = Field(default="SAR", max_length=8)
    image_url: Optional[str] = ""

    @model_validator(mode="after")
    def _require_identifier(self):
        if not (self.sku and self.sku.strip()) and not (
                self.product_id and self.product_id.strip()):
            raise ValueError(
                "يجب توفير SKU أو رقم المنتج (product_id) — أحدهما على الأقل."
            )
        return self


class ProductCostUpdate(BaseModel):
    sku: Optional[str] = None
    product_id: Optional[str] = None
    product_name: Optional[str] = None
    supplier_name: Optional[str] = None
    supplier_country: Optional[str] = None
    supplier_notes: Optional[str] = None
    cost_price: Optional[float] = Field(default=None, ge=0)
    currency: Optional[str] = None
    is_active: Optional[bool] = None
    image_url: Optional[str] = None
    cost_pending: Optional[bool] = None


# ── Router factory ─────────────────────────────────────────────────────────
def _build_router(db, current_user_dep) -> APIRouter:
    router = APIRouter(prefix="/product-costs", tags=["product-costs"])

    # ── List with search/filter ──────────────────────────────────────────
    @router.get("/")
    async def list_costs(
        search: Optional[str] = Query(None, description="Match name/sku/supplier"),
        is_active: Optional[bool] = Query(None),
        supplier: Optional[str] = Query(None),
        limit: int = Query(200, ge=1, le=1000),
        skip: int = Query(0, ge=0),
        user: dict = Depends(current_user_dep),
    ):
        q: dict = {"user_id": user["id"]}
        if is_active is not None:
            q["is_active"] = is_active
        if supplier:
            q["supplier_name"] = supplier
        if search:
            rx = re.escape(search.strip())
            q["$or"] = [
                {"product_name": {"$regex": rx, "$options": "i"}},
                {"sku": {"$regex": rx, "$options": "i"}},
                {"supplier_name": {"$regex": rx, "$options": "i"}},
            ]
        total = await db.product_costs.count_documents(q)
        rows = await db.product_costs.find(q, {"_id": 0}).sort(
            "updated_at", -1,
        ).skip(skip).limit(limit).to_list(limit)
        return {"items": rows, "total": total, "limit": limit, "skip": skip}

    @router.post("/")
    async def create_cost(payload: ProductCostIn, user: dict = Depends(current_user_dep)):
        uid = user["id"]
        # Iteration 25: SKU is optional. When absent, product_id is the
        # primary identifier and the row gets a synthetic sku_normalized
        # so the unique index still holds.
        sku = (payload.sku or "").strip()
        product_id = _norm_product_id(payload.product_id)
        effective_key = sku or product_id  # at least one is present (validator)
        sku_norm = _norm_sku(effective_key)
        # Iteration 25: cost_pending = True iff cost_price was not supplied.
        cost_pending = (payload.cost_price is None)
        cost_price = round(float(payload.cost_price or 0), 2)
        # Iteration 25: lookup order is product_id FIRST, then sku_normalized.
        # This preserves identity across re-imports where SKU may flip from
        # empty → real, since the product_id remains stable.
        existing = None
        if product_id:
            existing = await db.product_costs.find_one(
                {"user_id": uid, "product_id": product_id},
                {"_id": 0, "id": 1, "is_active": 1, "sku_normalized": 1},
            )
        if not existing:
            existing = await db.product_costs.find_one(
                {"user_id": uid, "sku_normalized": sku_norm},
                {"_id": 0, "id": 1, "is_active": 1, "sku_normalized": 1},
            )
        if existing and existing.get("is_active", True):
            raise HTTPException(
                status_code=409,
                detail=(f"المنتج موجود مسبقاً ({sku or product_id}). "
                        "عدّل القائمة الموجودة بدلاً من إضافة جديد."),
            )
        now = _now_iso()
        doc = {
            "id": str(uuid.uuid4()),
            "user_id": uid,
            "sku": sku,
            "sku_normalized": sku_norm,
            "product_id": product_id,
            "product_name": payload.product_name.strip(),
            "supplier_name": (payload.supplier_name or "").strip(),
            "supplier_country": (payload.supplier_country or "").strip(),
            "supplier_notes": (payload.supplier_notes or "").strip(),
            "cost_price": cost_price,
            "cost_pending": cost_pending,
            "currency": (payload.currency or "SAR").upper(),
            "image_url": (payload.image_url or "").strip(),
            "is_active": True,
            "meta": {},
            "created_at": now,
            "updated_at": now,
        }
        if existing:
            # Re-activate the soft-deleted row instead of creating a duplicate
            await db.product_costs.update_one(
                {"user_id": uid, "id": existing["id"]},
                {"$set": {**{k: v for k, v in doc.items()
                             if k not in ("id", "created_at")},
                          "is_active": True}},
            )
            doc = await db.product_costs.find_one(
                {"user_id": uid, "id": existing["id"]}, {"_id": 0},
            )
            doc["reprocessed_orders"] = await _reprocess_orders_for_keys(
                db, uid, {sku_norm}, {product_id},
            )
            # Iteration 26: also recompute last 2 days unconditionally
            # so the dashboard/reports refresh immediately.
            doc["recent_orders_recomputed"] = await _recompute_recent_orders(
                db, uid, days=2,
            )
            return doc
        await db.product_costs.insert_one(doc)
        doc.pop("_id", None)
        # Iteration 24/25/26: targeted reprocess (all-time for this SKU)
        # + recompute last 2 days (catches freshly-ingested orders).
        if not cost_pending:
            doc["reprocessed_orders"] = await _reprocess_orders_for_keys(
                db, uid, {sku_norm}, {product_id},
            )
        else:
            doc["reprocessed_orders"] = 0
        doc["recent_orders_recomputed"] = await _recompute_recent_orders(
            db, uid, days=2,
        )
        return doc

    @router.put("/{item_id}")
    async def update_cost(item_id: str, payload: ProductCostUpdate,
                          user: dict = Depends(current_user_dep)):
        patch = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
        if "product_name" in patch:
            patch["product_name"] = str(patch["product_name"]).strip()
        if "supplier_name" in patch:
            patch["supplier_name"] = str(patch["supplier_name"]).strip()
        if "supplier_country" in patch:
            patch["supplier_country"] = str(patch["supplier_country"]).strip()
        if "supplier_notes" in patch:
            patch["supplier_notes"] = str(patch["supplier_notes"]).strip()
        if "currency" in patch:
            patch["currency"] = str(patch["currency"]).upper()
        if "cost_price" in patch:
            patch["cost_price"] = round(float(patch["cost_price"]), 2)
            # Iteration 25: editing cost_price automatically clears the
            # cost_pending flag — the merchant has set a real price.
            patch["cost_pending"] = False
        if "sku" in patch:
            patch["sku"] = str(patch["sku"] or "").strip()
            patch["sku_normalized"] = _norm_sku(
                patch["sku"] or patch.get("product_id") or "")
        if "product_id" in patch:
            patch["product_id"] = _norm_product_id(patch["product_id"])
        if "image_url" in patch:
            patch["image_url"] = str(patch["image_url"] or "").strip()
        patch["updated_at"] = _now_iso()
        r = await db.product_costs.update_one(
            {"user_id": user["id"], "id": item_id}, {"$set": patch},
        )
        if r.matched_count == 0:
            raise HTTPException(status_code=404, detail="Product cost not found")
        doc = await db.product_costs.find_one(
            {"user_id": user["id"], "id": item_id}, {"_id": 0},
        )
        # Iteration 24/25: targeted reprocess — only when fields that affect
        # cost calculations changed (price / product_id / is_active /
        # cost_pending flip). Touching cost_pending=False is identical
        # to "cost just became available" so we ALWAYS reprocess in that case.
        if any(k in patch for k in ("cost_price", "product_id",
                                     "is_active", "cost_pending", "sku")):
            doc["reprocessed_orders"] = await _reprocess_orders_for_keys(
                db, user["id"],
                {_norm_sku(doc.get("sku") or doc.get("product_id") or "")},
                {_norm_product_id(doc.get("product_id"))},
            )
            # Iteration 26: also recompute last 2 days unconditionally.
            doc["recent_orders_recomputed"] = await _recompute_recent_orders(
                db, user["id"], days=2,
            )
        return doc

    @router.delete("/{item_id}")
    async def delete_cost(item_id: str, user: dict = Depends(current_user_dep)):
        r = await db.product_costs.update_one(
            {"user_id": user["id"], "id": item_id},
            {"$set": {"is_active": False, "updated_at": _now_iso()}},
        )
        if r.matched_count == 0:
            raise HTTPException(status_code=404, detail="Product cost not found")
        return {"ok": True}

    # ── Excel import ─────────────────────────────────────────────────────
    # Per merchant decision (iteration 20):
    #   • Excel imports ONLY 3 columns into the cost calculation:
    #       SKU, product_name, cost_price.
    #   • EVERY OTHER COLUMN in the merchant's Salla export is preserved
    #     in `meta` (a free-form dict) so we don't lose data and can
    #     surface it later (filters, profitability reports, etc) — but
    #     `meta` MUST NEVER influence financial calculations.
    #   • Supplier (name/country/notes) is EXCLUSIVELY managed inside the
    #     UI (Add/Edit modal) — NOT imported, NOT inferred from Excel.
    HEADER_ALIASES = {
        # SKU — explicit SKU/Reference/Product Code. Note: "رقم المنتج" is
        # ambiguous in Arabic — Salla uses it for the numeric product_id,
        # NOT for the SKU. So it lives under product_id below (iteration 22).
        "sku": {
            "sku", "كود المنتج", "كود", "الرمز",
            "reference", "product code", "code", "product_code",
            "item code", "item_code", "barcode-sku", "merchant_sku",
        },
        # product name
        "product_name": {
            "product_name", "name", "اسم المنتج", "الاسم", "المنتج",
            "title", "product title", "product",
        },
        # cost — every Arabic / English variant
        "cost_price": {
            "cost", "cost_price", "purchase_price", "purchase price",
            "buy price", "buy_price", "price_cost",
            "التكلفة", "تكلفة الشراء", "تكلفة المنتج",
            "سعر التكلفة", "سعر الشراء",
            "الكلفة", "كلفة المنتج",
        },
        # product_id (Salla's internal numeric ID — every common alias).
        # When the merchant's Salla export omits SKU, this becomes the
        # primary identifier (iteration 22 — falls back gracefully).
        "product_id": {
            "product_id", "id", "معرف المنتج",
            "رقم المنتج", "product id", "product-id",
        },
        # currency (rarely present, defaults SAR)
        "currency": {"currency", "العملة"},
        # image URL (iteration 23) — Salla product Excel exports place the
        # image URL in column F by default. If a header alias is missing
        # we fall back to column index 5 (F) at parse time.
        "image_url": {
            "image", "image_url", "image url", "img", "img_url",
            "صورة", "صورة المنتج", "الصورة", "رابط الصورة",
            "photo", "picture", "thumbnail",
        },
    }
    # Mapped header keys we KNOW about — every other header is preserved in meta.
    _MAPPED_HEADERS = (
        HEADER_ALIASES["sku"]
        | HEADER_ALIASES["product_name"]
        | HEADER_ALIASES["cost_price"]
        | HEADER_ALIASES["product_id"]
        | HEADER_ALIASES["currency"]
        | HEADER_ALIASES["image_url"]
    )

    def _find_col(headers: list, target: str) -> Optional[int]:
        targets = {h.lower() for h in HEADER_ALIASES[target]}
        for i, h in enumerate(headers):
            if str(h or "").strip().lower() in targets:
                return i
        return None

    @router.post("/import")
    async def import_excel(
        file: UploadFile = File(...),
        update_existing: bool = Query(
            True,
            description="If True (default), rows with an existing SKU are "
                        "UPDATED (price / name overwritten). If False, "
                        "duplicate SKUs are SKIPPED and reported under "
                        "`skipped`. Maps to the UI checkbox 'تحديث "
                        "المنتجات الموجودة بنفس SKU'.",
        ),
        user: dict = Depends(current_user_dep),
    ):
        if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400,
                                detail="الملف يجب أن يكون Excel (.xlsx أو .xls)")
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(status_code=500,
                                detail="openpyxl missing — install it on the backend")
        content = await file.read()
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400,
                                detail=f"تعذر قراءة الملف: {exc}")
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="الملف فارغ")
        headers_raw = [str(c or "").strip() for c in rows[0]]
        headers_lc = [h.lower() for h in headers_raw]
        idx_sku = _find_col(headers_raw, "sku")
        idx_name = _find_col(headers_raw, "product_name")
        idx_cost = _find_col(headers_raw, "cost_price")
        idx_product_id = _find_col(headers_raw, "product_id")
        idx_currency = _find_col(headers_raw, "currency")
        # Iteration 23: image URL detection. Try header aliases first, then
        # fall back to column F (index 5) which is the default position
        # for product image URLs in Salla's product Excel export.
        # Guard 1: only use the fallback if column F isn't already claimed
        # by another mapped column (sku/name/cost/product_id/currency).
        # Guard 2: only use the fallback if ≥1 data row in column F contains
        # a URL-looking value. Otherwise treat F as a regular meta column
        # so we don't accidentally swallow unrelated data (e.g. category).
        idx_image = _find_col(headers_raw, "image_url")
        if idx_image is None and len(headers_raw) > 5:
            used = {idx_sku, idx_name, idx_cost, idx_product_id, idx_currency}
            if 5 not in used:
                def _looks_like_url(v: Any) -> bool:
                    s = str(v or "").strip()
                    if not s:
                        return False
                    return s.startswith(("http://", "https://", "//")) or any(
                        s.lower().endswith(ext) for ext in
                        (".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg")
                    )
                # Scan up to first 200 data rows for performance on big files.
                for row in rows[1:201]:
                    if len(row) > 5 and _looks_like_url(row[5]):
                        idx_image = 5  # column F → image
                        break
        # Iteration 22+25: only SKU OR product_id is mandatory. Cost is
        # also OPTIONAL — rows with empty cost are saved as "pending
        # cost" (cost_pending=True) so the merchant can see them in the
        # catalogue immediately and fill prices later. We NEVER treat
        # missing cost as 0.
        if idx_sku is None and idx_product_id is None:
            raise HTTPException(
                status_code=400,
                detail="الأعمدة المطلوبة: SKU أو رقم المنتج (أحدهما على الأقل). "
                       "التكلفة واسم المنتج اختياريان. باقي الأعمدة تُحفظ في meta.",
            )

        # Compute meta-column indices (every header that's NOT one of our
        # mapped headers). These get saved into doc.meta verbatim.
        # Iteration 23: also exclude the image column when it was matched
        # via the column-F fallback (its header isn't in _MAPPED_HEADERS
        # because we matched by position, not by name).
        meta_cols: list[tuple[int, str]] = []
        for i, h in enumerate(headers_lc):
            if h and h not in _MAPPED_HEADERS and i != idx_image:
                meta_cols.append((i, headers_raw[i]))

        uid = user["id"]
        now = _now_iso()
        created = 0
        updated = 0
        skipped = 0
        images_imported = 0
        pending_count = 0  # iteration 25: rows imported without cost
        # Iteration 25: collect every (sku_norm, product_id) that landed
        # with a REAL cost so we can fire a single targeted reprocess
        # after the loop ends. Pending rows are excluded — they don't
        # contribute any cost yet, so reprocessing them is a no-op.
        reprocess_skus: set = set()
        reprocess_pids: set = set()
        errors: list[dict] = []
        for row_num, r in enumerate(rows[1:], start=2):
            try:
                sku = (str(r[idx_sku] or "").strip()
                       if (idx_sku is not None and idx_sku < len(r)) else "")
                product_id = (str(r[idx_product_id] or "").strip()
                              if (idx_product_id is not None and idx_product_id < len(r))
                              else "")
                # Iteration 22: require ONE of sku OR product_id per row.
                if not sku and not product_id:
                    continue
                # Iteration 25: PRODUCT_ID is the primary identifier when
                # present (Salla product Excel exports almost always have
                # `رقم المنتج` but not always SKU). SKU acts as fallback.
                product_id_norm = _norm_product_id(product_id)
                effective_key = product_id_norm or sku
                name = (str(r[idx_name] or "").strip()
                        if (idx_name is not None and idx_name < len(r)) else "")
                if not name:
                    name = effective_key
                # Iteration 25: cost is OPTIONAL per row. Empty/None cost
                # cells are saved with cost_pending=True, NOT treated as 0.
                raw_cost = (r[idx_cost] if (idx_cost is not None and idx_cost < len(r))
                            else None)
                cost_provided = (raw_cost is not None
                                 and str(raw_cost).strip() != "")
                cost = _to_float(raw_cost) if cost_provided else 0.0
                if cost < 0:
                    errors.append({"row": row_num, "error": "التكلفة سالبة"})
                    continue
                cost_pending = not cost_provided
                currency = ((str(r[idx_currency] or "SAR").strip().upper() or "SAR")
                            if (idx_currency is not None and idx_currency < len(r))
                            else "SAR")
                # Iteration 23: parse image URL from the resolved column.
                image_url = ""
                if idx_image is not None and idx_image < len(r):
                    raw_img = str(r[idx_image] or "").strip()
                    if raw_img and (
                        raw_img.startswith(("http://", "https://", "//", "/"))
                        or any(raw_img.lower().endswith(ext) for ext in (
                            ".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg",
                        ))
                    ):
                        image_url = raw_img
                meta: dict = {}
                for col_idx, col_label in meta_cols:
                    if col_idx < len(r):
                        cell = r[col_idx]
                        if cell is not None and str(cell).strip() != "":
                            meta[col_label] = cell
                sku_norm = _norm_sku(effective_key)

                # Iteration 25: prefer PRODUCT_ID as the upsert key when
                # available — keeps merchant's catalogue de-duplicated
                # across re-imports where SKU may flip from empty → real
                # (or vice-versa). Falls back to sku_normalized only for
                # legacy/non-Salla rows that have ONLY SKU.
                find_query: dict
                if product_id_norm:
                    find_query = {"user_id": uid, "product_id": product_id_norm}
                else:
                    find_query = {"user_id": uid, "sku_normalized": sku_norm}

                # Apply update_existing flag — skip rows whose identity
                # is already present when the merchant un-checked the box.
                if not update_existing:
                    existing = await db.product_costs.find_one(
                        find_query, {"_id": 0, "id": 1},
                    )
                    if existing:
                        skipped += 1
                        continue

                doc = {
                    "user_id": uid,
                    "sku": sku,  # may be empty when only product_id was provided
                    "sku_normalized": sku_norm,
                    "product_id": product_id_norm,
                    "product_name": name,
                    # NOTE: supplier_* fields are PRESERVED — we never
                    # touch them on import (manual UI management only).
                    "cost_price": round(cost, 2),
                    "cost_pending": cost_pending,
                    "currency": currency,
                    "is_active": True,
                    "meta": meta,
                    "updated_at": now,
                }
                # Iteration 23: only $set image_url when this row has one
                # — preserves any manually-uploaded image on re-imports
                # whose Excel doesn't include the image column.
                if image_url:
                    doc["image_url"] = image_url
                res = await db.product_costs.update_one(
                    find_query,
                    {"$set": doc,
                     "$setOnInsert": {
                         "id": str(uuid.uuid4()),
                         "supplier_name": "",
                         "supplier_country": "",
                         "supplier_notes": "",
                         # Default image_url on insert (kept empty so we
                         # don't override existing values via $set above).
                         **({"image_url": ""} if not image_url else {}),
                         "created_at": now,
                     }},
                    upsert=True,
                )
                if res.upserted_id:
                    created += 1
                else:
                    updated += 1
                if image_url:
                    images_imported += 1
                if cost_pending:
                    pending_count += 1
                else:
                    # Track keys to reprocess once at the end of the loop.
                    if sku_norm:
                        reprocess_skus.add(sku_norm)
                    if product_id_norm:
                        reprocess_pids.add(product_id_norm)
            except Exception as exc:
                errors.append({"row": row_num, "error": str(exc)[:200]})

        # Iteration 25: ONE targeted reprocess pass for all keys with real
        # cost — flips affected past orders from incomplete → complete.
        reprocessed_orders = 0
        if reprocess_skus or reprocess_pids:
            reprocessed_orders = await _reprocess_orders_for_keys(
                db, uid, reprocess_skus, reprocess_pids,
            )
        # Iteration 26: unconditional recompute of last 2 days so the
        # dashboard/reports reflect the latest costs immediately.
        recent_recomputed = await _recompute_recent_orders(db, uid, days=2)

        return {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "pending_count": pending_count,
            "errors": errors,
            "total_processed": created + updated,
            "update_existing": update_existing,
            "meta_columns_preserved": [c[1] for c in meta_cols],
            "images_imported": images_imported,
            "image_column_detected": (
                "header" if _find_col(headers_raw, "image_url") is not None
                else ("column_F" if idx_image == 5 else None)
            ),
            "reprocessed_orders": reprocessed_orders,
            "recent_orders_recomputed": recent_recomputed,
        }

    # ── Missing costs (orders with unmatched products) ───────────────────
    @router.get("/missing")
    async def missing_costs(
        days: int = Query(60, ge=1, le=365),
        user: dict = Depends(current_user_dep),
    ):
        """Aggregate every order line in the last `days` whose SKU/product_id
        has NO matching active product_costs entry. Returns counts + image
        + last-order info so the merchant can quickly add the missing
        cost from the "Missing Products" UI (iteration 24).

        Also returns a side-channel `excel_no_products_count` — the number
        of orders that came in WITHOUT a `products[]` array at all (most
        commonly Excel-imported orders). Those orders cannot be auto-cost
        matched until the merchant edits them manually.
        """
        uid = user["id"]
        cutoff = (datetime.now(timezone.utc).date() - timedelta(days=days)).isoformat()
        # Fetch recent orders' products + their missing-lines arrays.
        cursor = db.unified_orders.find(
            {"user_id": uid, "order_date": {"$gte": cutoff}},
            {"_id": 0, "missing_product_cost_lines": 1, "products": 1,
             "order_number": 1, "order_date": 1, "data_source": 1,
             "profit_status": 1},
        )
        agg: dict[str, dict] = {}
        excel_no_products = 0
        async for o in cursor:
            lines = o.get("missing_product_cost_lines")
            if lines is None:
                # Order has not been cost-enriched yet — compute lazily.
                _, _, missing = await compute_order_cost(
                    db, uid, o.get("products") or [],
                )
                lines = missing
            # Track Excel-without-products orders (iteration 24).
            if (o.get("profit_status") == "incomplete_no_products" or
                    not o.get("products")):
                if (o.get("data_source") or "").lower() in ("excel", ""):
                    excel_no_products += 1
            for ln in (lines or []):
                key = (_norm_sku(ln.get("sku")) or
                       _norm_product_id(ln.get("product_id")) or
                       (ln.get("name") or "").strip())
                if not key:
                    continue
                cur = agg.setdefault(key, {
                    "sku": ln.get("sku") or "",
                    "product_id": ln.get("product_id") or "",
                    "name": ln.get("name") or "",
                    "image_url": ln.get("image_url") or "",
                    "occurrences": 0,
                    "total_quantity": 0.0,
                    "last_order_number": "",
                    "last_order_date": "",
                })
                cur["occurrences"] += 1
                cur["total_quantity"] += _to_float(ln.get("quantity"), 1.0)
                # Track most recent order this missing line appeared in.
                od = o.get("order_date") or ""
                if od >= cur["last_order_date"]:
                    cur["last_order_date"] = od
                    cur["last_order_number"] = o.get("order_number") or ""
                # Fill image_url lazily from any order line that has one.
                if not cur["image_url"] and ln.get("image_url"):
                    cur["image_url"] = ln["image_url"]

        # Iteration 25: also include catalogue rows with cost_pending=True
        # (imported via Excel without a cost_price). Marked source="catalogue"
        # so the UI can show them even if no order arrived yet.
        async for cat in db.product_costs.find(
            {"user_id": uid, "is_active": True, "cost_pending": True},
            {"_id": 0, "sku": 1, "product_id": 1, "product_name": 1,
             "image_url": 1, "id": 1},
        ):
            key = (_norm_sku(cat.get("sku")) or
                   _norm_product_id(cat.get("product_id")) or
                   (cat.get("product_name") or "").strip())
            if not key:
                continue
            cur = agg.setdefault(key, {
                "sku": cat.get("sku") or "",
                "product_id": cat.get("product_id") or "",
                "name": cat.get("product_name") or "",
                "image_url": cat.get("image_url") or "",
                "occurrences": 0,
                "total_quantity": 0.0,
                "last_order_number": "",
                "last_order_date": "",
            })
            cur["pending_in_catalogue"] = True
            cur["catalogue_id"] = cat.get("id")
            # If catalogue has an image but the order didn't, fill it in.
            if not cur["image_url"] and cat.get("image_url"):
                cur["image_url"] = cat["image_url"]

        rows = sorted(agg.values(), key=lambda r: r["occurrences"], reverse=True)
        return {
            "items": rows,
            "count": len(rows),
            "window_days": days,
            "excel_no_products_count": excel_no_products,
        }

    # ── Summary stats ────────────────────────────────────────────────────
    @router.get("/summary")
    async def summary(user: dict = Depends(current_user_dep)):
        uid = user["id"]
        today_d = datetime.now(timezone.utc).date()
        today_str = today_d.isoformat()
        month_start = today_str[:8] + "01"
        # Iteration 27/28: lazy self-heal — before computing today/month
        # totals, re-attach cost on every order in the CURRENT MONTH
        # whose total_product_cost is still null. This is cheap (only
        # touches stale rows) and idempotent. It guarantees the Dashboard
        # card always reflects reality even on environments that don't
        # have the iteration-26 auto-recompute hooks.
        #
        # Iteration 28: widened from "today only" → "whole current month"
        # because the merchant reported month_total stayed 0 even after
        # today's orders healed (older-in-month orders were still stale).
        stale_today_healed = 0
        stale_month_healed = 0
        async for o in db.unified_orders.find(
            {"user_id": uid,
             "order_date": {"$gte": month_start, "$lte": today_str},
             "$or": [
                 {"total_product_cost": None},
                 {"total_product_cost": {"$exists": False}},
             ]},
            {"_id": 0, "order_number": 1, "products": 1, "order_date": 1},
        ):
            try:
                patch = await attach_cost_to_order_doc(db, uid, o)
                await db.unified_orders.update_one(
                    {"user_id": uid, "order_number": o["order_number"]},
                    {"$set": patch},
                )
                stale_month_healed += 1
                if o.get("order_date") == today_str:
                    stale_today_healed += 1
            except Exception:
                pass  # never fail summary if heal fails on one row
        # Today / month: sum total_product_cost on orders in that range.
        today_total = 0.0
        month_total = 0.0
        async for o in db.unified_orders.find(
            {"user_id": uid, "order_date": {"$gte": month_start, "$lte": today_str}},
            {"_id": 0, "order_date": 1, "total_product_cost": 1},
        ):
            cost = float(o.get("total_product_cost") or 0)
            if o.get("order_date") == today_str:
                today_total += cost
            month_total += cost
        # Avg per active product
        active = await db.product_costs.count_documents(
            {"user_id": uid, "is_active": True},
        )
        # Iteration 26: split active count into "linked" (cost set) and
        # "missing" (cost_pending). Surfaces directly on the Dashboard
        # product-cost card.
        linked_products_count = await db.product_costs.count_documents(
            {"user_id": uid, "is_active": True, "cost_pending": {"$ne": True}},
        )
        catalogue_pending_count = await db.product_costs.count_documents(
            {"user_id": uid, "is_active": True, "cost_pending": True},
        )
        # Also count distinct missing-cost SKUs across recent orders that
        # aren't in the catalogue yet → "products ordered but never seeded".
        cutoff_60 = (today_d - timedelta(days=59)).isoformat()
        order_missing_keys: set = set()
        async for o in db.unified_orders.find(
            {"user_id": uid, "order_date": {"$gte": cutoff_60}},
            {"_id": 0, "missing_product_cost_lines": 1},
        ):
            for ln in (o.get("missing_product_cost_lines") or []):
                k = (_norm_sku(ln.get("sku")) or
                     _norm_product_id(ln.get("product_id")))
                if k:
                    order_missing_keys.add(k)
        # Some of those keys ARE in the catalogue as cost_pending → already
        # counted. Subtract the overlap so we don't double-count.
        if order_missing_keys:
            existing_keys = set()
            async for c in db.product_costs.find(
                {"user_id": uid, "is_active": True, "cost_pending": True,
                 "$or": [
                     {"sku_normalized": {"$in": list(order_missing_keys)}},
                     {"product_id": {"$in": list(order_missing_keys)}},
                 ]},
                {"_id": 0, "sku_normalized": 1, "product_id": 1},
            ):
                if c.get("sku_normalized"):
                    existing_keys.add(c["sku_normalized"])
                if c.get("product_id"):
                    existing_keys.add(c["product_id"])
            order_only_missing = order_missing_keys - existing_keys
        else:
            order_only_missing = set()
        missing_products_count = catalogue_pending_count + len(order_only_missing)

        if active:
            agg = db.product_costs.aggregate([
                {"$match": {"user_id": uid, "is_active": True,
                            "cost_pending": {"$ne": True}}},
                {"$group": {"_id": None, "avg": {"$avg": "$cost_price"}}},
            ])
            r = await agg.to_list(1)
            avg_cost = round(float(r[0]["avg"]) if r else 0.0, 2)
        else:
            avg_cost = 0.0
        # Top profitable products in last 30 days (by line_cost in cost_items)
        cutoff_30 = (today_d - timedelta(days=29)).isoformat()
        rev_by_sku: dict[str, dict] = {}
        async for o in db.unified_orders.find(
            {"user_id": uid, "order_date": {"$gte": cutoff_30, "$lte": today_str}},
            {"_id": 0, "cost_items": 1, "products": 1, "total_amount": 1},
        ):
            items = o.get("cost_items") or []
            for it in items:
                sku = _norm_sku(it.get("sku")) or _norm_product_id(it.get("product_id"))
                if not sku:
                    continue
                cur = rev_by_sku.setdefault(sku, {
                    "sku": it.get("sku") or "",
                    "product_id": it.get("product_id") or "",
                    "name": it.get("name") or "",
                    "qty_sold": 0.0,
                    "total_cost": 0.0,
                })
                cur["qty_sold"] += _to_float(it.get("quantity"), 1.0)
                cur["total_cost"] += _to_float(it.get("line_cost"), 0.0)
        top = sorted(rev_by_sku.values(), key=lambda r: r["total_cost"], reverse=True)[:10]
        return {
            "today_total": round(today_total, 2),
            "month_total": round(month_total, 2),
            "month_start": month_start,
            "active_products": active,
            "linked_products_count": int(linked_products_count),
            "missing_products_count": int(missing_products_count),
            "avg_cost": avg_cost,
            "top_products_last_30d": top,
            "stale_today_healed": int(stale_today_healed),
            "stale_month_healed": int(stale_month_healed),
            "currency": "SAR",
        }

    # ── Product Sales Report (iteration 26) ──────────────────────────────
    @router.get("/product-sales")
    async def product_sales_report(
        from_date: Optional[str] = Query(None),
        to_date: Optional[str] = Query(None),
        user: dict = Depends(current_user_dep),
    ):
        """Aggregate orders by product → return per-product breakdown for
        the merchant's "Product Sales Report".

        Returns: image_url, name, product_id, sku, units_sold,
        total_sales, total_cost, total_profit, profit_margin_pct,
        cost_status ('complete' or 'incomplete').

        Defaults to the last 2 days when no date range is supplied, per
        merchant requirement: "أعتمد على بيانات آخر يومين على الأقل".
        """
        uid = user["id"]
        today_d = datetime.now(timezone.utc).date()
        if not to_date:
            to_date = today_d.isoformat()
        if not from_date:
            from_date = (today_d - timedelta(days=1)).isoformat()

        # Pre-load catalogue (for image_url + cost_pending lookup).
        cat_by_sku: dict[str, dict] = {}
        cat_by_pid: dict[str, dict] = {}
        async for c in db.product_costs.find(
            {"user_id": uid, "is_active": True},
            {"_id": 0, "sku_normalized": 1, "product_id": 1,
             "image_url": 1, "product_name": 1, "cost_pending": 1,
             "cost_price": 1, "sku": 1},
        ):
            if c.get("sku_normalized"):
                cat_by_sku[c["sku_normalized"]] = c
            if c.get("product_id"):
                cat_by_pid[c["product_id"]] = c

        agg: dict[str, dict] = {}
        async for o in db.unified_orders.find(
            {"user_id": uid,
             "order_date": {"$gte": from_date, "$lte": to_date}},
            {"_id": 0, "products": 1, "cost_items": 1,
             "missing_product_cost_lines": 1, "total_amount": 1,
             "total": 1, "order_date": 1, "order_number": 1,
             "profit_status": 1},
        ):
            products = o.get("products") or []
            cost_items = {((_norm_sku(it.get("sku")) or
                            _norm_product_id(it.get("product_id")))
                           or (it.get("name") or "")): it
                          for it in (o.get("cost_items") or [])}
            for p in products:
                sku = _norm_sku(p.get("sku"))
                pid = _norm_product_id(p.get("product_id") or p.get("id"))
                name = (p.get("name") or "").strip()
                key = sku or pid or name
                if not key:
                    continue
                qty = _to_float(p.get("quantity"), 1.0)
                price = _to_float(p.get("price"), 0.0)
                line_sales = round(price * qty, 2)
                # Match cost.
                ci = cost_items.get(sku) or cost_items.get(pid) or cost_items.get(name)
                line_cost = _to_float(ci.get("line_cost") if ci else 0, 0.0)
                matched = bool(ci and ci.get("matched_by"))
                # Image: cost_items > catalogue > webhook.
                image_url = ""
                cat = (cat_by_sku.get(sku) if sku else None) or \
                      (cat_by_pid.get(pid) if pid else None)
                if cat and cat.get("image_url"):
                    image_url = cat["image_url"]
                if not image_url:
                    image_url = (p.get("image_url") or p.get("image") or "").strip()
                cur = agg.setdefault(key, {
                    "product_id": pid or "",
                    "sku": p.get("sku") or "",  # original casing for display
                    "name": name,
                    "image_url": image_url,
                    "units_sold": 0.0,
                    "total_sales": 0.0,
                    "total_cost": 0.0,
                    "matched_units": 0.0,
                    "currency": (cat or {}).get("currency", "SAR"),
                })
                cur["units_sold"] += qty
                cur["total_sales"] += line_sales
                cur["total_cost"] += line_cost
                if matched:
                    cur["matched_units"] += qty
                if not cur["image_url"] and image_url:
                    cur["image_url"] = image_url

        rows: list[dict] = []
        for r in agg.values():
            sold = r["units_sold"]
            matched = r["matched_units"]
            complete = sold > 0 and matched >= sold - 0.0001
            profit = round(r["total_sales"] - r["total_cost"], 2) if complete else None
            margin = (round((profit / r["total_sales"]) * 100, 2)
                      if (complete and r["total_sales"] > 0) else None)
            rows.append({
                "product_id": r["product_id"],
                "sku": r["sku"],
                "name": r["name"],
                "image_url": r["image_url"],
                "units_sold": round(sold, 2),
                "total_sales": round(r["total_sales"], 2),
                "total_cost": round(r["total_cost"], 2),
                "total_profit": profit,
                "profit_margin_pct": margin,
                "cost_status": "complete" if complete else "incomplete",
                "currency": r["currency"],
            })

        # Totals (only over rows with COMPLETE cost data — per merchant rule).
        total_sales_all = round(sum(r["total_sales"] for r in rows), 2)
        complete_rows = [r for r in rows if r["cost_status"] == "complete"]
        incomplete_rows = [r for r in rows if r["cost_status"] != "complete"]
        total_sales_complete = round(sum(r["total_sales"] for r in complete_rows), 2)
        total_cost_complete = round(sum(r["total_cost"] for r in complete_rows), 2)
        total_profit_complete = round(total_sales_complete - total_cost_complete, 2)
        margin_complete = (round((total_profit_complete / total_sales_complete) * 100, 2)
                           if total_sales_complete > 0 else 0.0)
        # Sort: incomplete first (so merchant notices them), then by sales desc.
        rows.sort(key=lambda r: (r["cost_status"] == "complete",
                                 -r["total_sales"]))
        return {
            "range": {"from_date": from_date, "to_date": to_date},
            "items": rows,
            "count": len(rows),
            "incomplete_count": len(incomplete_rows),
            "totals": {
                "total_sales_all": total_sales_all,
                "total_sales_complete": total_sales_complete,
                "total_cost_complete": total_cost_complete,
                "total_profit_complete": total_profit_complete,
                "margin_complete_pct": margin_complete,
            },
            "currency": "SAR",
        }

    # ── Recompute (after import or bulk edits) ───────────────────────────
    @router.post("/recompute")
    async def recompute_orders(
        days: int = Query(60, ge=1, le=365),
        user: dict = Depends(current_user_dep),
    ):
        """Re-attach cost data to every order in the last `days`. Useful
        after a bulk Excel import — orders ingested BEFORE the cost was
        known will now reflect the new cost."""
        uid = user["id"]
        cutoff = (datetime.now(timezone.utc).date() - timedelta(days=days)).isoformat()
        updated = 0
        async for o in db.unified_orders.find(
            {"user_id": uid, "order_date": {"$gte": cutoff}},
            {"_id": 0, "order_number": 1, "products": 1},
        ):
            patch = await attach_cost_to_order_doc(db, uid, o)
            await db.unified_orders.update_one(
                {"user_id": uid, "order_number": o["order_number"]},
                {"$set": patch},
            )
            updated += 1
        return {"orders_updated": updated, "window_days": days}

    return router


def attach_product_costs_routes(parent_router, db, current_user_dep) -> None:
    parent_router.include_router(_build_router(db, current_user_dep))
