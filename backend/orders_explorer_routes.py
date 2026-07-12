"""Orders explorer (Iter-85 + Iter-86 export)

Endpoints:
  GET /api/orders                  → paginated list with filters
  GET /api/orders/status-summary   → historical counts + totals per status
  GET /api/orders/export.xlsx      → Excel export honouring same filters
"""

from io import BytesIO
import logging
from typing import Optional
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse

from auth import get_current_user_from_db
from order_status_policy import default_category_for, get_policy_map, resolve_category


logger = logging.getLogger(__name__)


def attach_orders_explorer_routes(parent_router: APIRouter, db) -> None:
    router = APIRouter(prefix="/orders", tags=["orders"])

    async def current_user(request: Request) -> dict:
        return await get_current_user_from_db(request, db)

    @router.get("/status-summary")
    async def status_summary(
        from_date: Optional[str] = Query(None),
        to_date: Optional[str] = Query(None),
        user: dict = Depends(current_user),
    ):
        """Aggregate counts + totals per order_status. When date filters
        are passed, returns the slice for that range; otherwise full
        historical data. Always includes the policy category so the UI
        can colour the chip."""
        uid = user["id"]
        match: dict = {"user_id": uid}
        if from_date or to_date:
            match["order_date"] = {}
            if from_date:
                match["order_date"]["$gte"] = from_date
            if to_date:
                match["order_date"]["$lte"] = to_date

        overrides = await get_policy_map(db, uid)

        pipeline = [
            {"$match": match},
            {"$group": {
                "_id": {"$ifNull": ["$order_status", ""]},
                "orders_count": {"$sum": 1},
                "total_amount": {"$sum": {"$ifNull": ["$total_amount", 0]}},
            }},
            {"$sort": {"orders_count": -1}},
        ]
        rows = []
        grand_count, grand_total = 0, 0.0
        async for r in db.unified_orders.aggregate(pipeline):
            raw = (r["_id"] or "").strip()
            display = raw if raw and raw != "\\N" else "(بدون حالة)"
            category = resolve_category(
                raw if raw and raw != "\\N" else None, overrides
            )
            rows.append({
                "status": raw or "__none__",
                "display": display,
                "orders_count": int(r["orders_count"]),
                "total_amount": round(float(r["total_amount"]) or 0, 2),
                "category": category,
                "default_category": default_category_for(
                    raw if raw and raw != "\\N" else None
                ),
            })
            grand_count += int(r["orders_count"])
            grand_total += float(r["total_amount"]) or 0

        by_category = {"confirmed": {"count": 0, "amount": 0.0},
                       "pending":   {"count": 0, "amount": 0.0},
                       "refunded":  {"count": 0, "amount": 0.0},
                       "cancelled": {"count": 0, "amount": 0.0}}
        for r in rows:
            c = r["category"]
            if c in by_category:
                by_category[c]["count"] += r["orders_count"]
                by_category[c]["amount"] += r["total_amount"]
        for c in by_category:
            by_category[c]["amount"] = round(by_category[c]["amount"], 2)

        return {
            "rows": rows,
            "by_category": by_category,
            "totals": {
                "orders_count": grand_count,
                "total_amount": round(grand_total, 2),
            },
        }

    async def _build_query(
        uid: str,
        from_date: Optional[str],
        to_date: Optional[str],
        status: Optional[str],
        category: Optional[str],
        payment_method: Optional[str],
        search: Optional[str],
        overrides: dict,
    ) -> dict:
        """Shared Mongo query builder used by list + export."""
        q: dict = {"user_id": uid}
        and_clauses: list[dict] = []

        if from_date or to_date:
            q["order_date"] = {}
            if from_date:
                q["order_date"]["$gte"] = from_date
            if to_date:
                q["order_date"]["$lte"] = to_date

        if status:
            if status == "__none__":
                and_clauses.append({"$or": [
                    {"order_status": {"$in": [None, ""]}},
                    {"order_status": "\\N"},
                ]})
            else:
                q["order_status"] = status

        if category:
            distinct_statuses = await db.unified_orders.distinct(
                "order_status", {"user_id": uid}
            )
            matching: list = []
            include_none = False
            for s in distinct_statuses:
                cat = resolve_category(s, overrides)
                if cat == category:
                    if s in (None, "", "\\N"):
                        include_none = True
                    else:
                        matching.append(s)
            if include_none:
                and_clauses.append({"$or": [
                    {"order_status": {"$in": matching + [None, "", "\\N"]}},
                ]})
            else:
                q["order_status"] = {"$in": matching or ["__none_match__"]}

        if payment_method:
            q["payment_method"] = payment_method

        if search:
            s = search.strip()
            and_clauses.append({"$or": [
                {"order_number": {"$regex": s, "$options": "i"}},
                {"customer_name": {"$regex": s, "$options": "i"}},
                {"customer_phone": {"$regex": s, "$options": "i"}},
            ]})

        if and_clauses:
            q["$and"] = and_clauses
        return q

    @router.get("")
    async def list_orders(
        from_date: Optional[str] = Query(None),
        to_date: Optional[str] = Query(None),
        status: Optional[str] = Query(None, description="exact order_status or __none__"),
        category: Optional[str] = Query(None, description="policy category filter"),
        payment_method: Optional[str] = Query(None),
        search: Optional[str] = Query(None, description="order_number / customer_name / customer_phone"),
        page: int = Query(1, ge=1),
        limit: int = Query(50, ge=1, le=500),
        user: dict = Depends(current_user),
    ):
        uid = user["id"]
        overrides = await get_policy_map(db, uid)
        q = await _build_query(uid, from_date, to_date, status, category,
                               payment_method, search, overrides)

        skip = (page - 1) * limit
        total = await db.unified_orders.count_documents(q)
        cursor = (
            db.unified_orders.find(q, {
                "_id": 0,
                "raw_by_source": 0,
                "raw_by_user": 0,
            })
            .sort("order_date", -1)
            .skip(skip)
            .limit(limit)
        )
        items = []
        async for o in cursor:
            o["category"] = resolve_category(o.get("order_status"), overrides)
            items.append(o)

        return {
            "items": items,
            "page": page,
            "limit": limit,
            "total": total,
            "pages": (total + limit - 1) // limit if total > 0 else 1,
        }

    CATEGORY_LABEL_AR = {
        "confirmed": "مؤكدة",
        "pending":   "معلّقة",
        "refunded":  "مسترجعة",
        "cancelled": "ملغاة",
    }

    @router.get("/export.xlsx")
    async def export_orders_xlsx(
        from_date: Optional[str] = Query(None),
        to_date: Optional[str] = Query(None),
        status: Optional[str] = Query(None),
        category: Optional[str] = Query(None),
        payment_method: Optional[str] = Query(None),
        search: Optional[str] = Query(None),
        user: dict = Depends(current_user),
    ):
        """Iter-86 — Excel export of the currently-filtered orders."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from datetime import datetime, timezone

        uid = user["id"]
        overrides = await get_policy_map(db, uid)
        q = await _build_query(uid, from_date, to_date, status, category,
                               payment_method, search, overrides)

        cursor = db.unified_orders.find(q, {
            "_id": 0,
            "raw_by_source": 0,
            "raw_by_user": 0,
        }).sort("order_date", -1).limit(50_000)

        rows: list[dict] = []
        async for o in cursor:
            o["category"] = resolve_category(o.get("order_status"), overrides)
            rows.append(o)

        wb = Workbook()
        ws = wb.active
        ws.title = "الطلبات"
        ws.sheet_view.rightToLeft = True

        headers = [
            "رقم الطلب", "تاريخ الطلب", "اسم العميل", "جوال العميل",
            "طريقة الدفع", "شركة الشحن", "الإجمالي (ر.س)",
            "حالة الطلب", "الفئة", "المدينة", "ملاحظات",
        ]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="0F5D46")
        header_font = Font(bold=True, color="FFFFFF", name="Tajawal")
        for col_idx, _ in enumerate(headers, 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        for o in rows:
            ws.append([
                o.get("order_number") or "",
                o.get("order_date") or "",
                o.get("customer_name") or "",
                o.get("customer_phone") or "",
                o.get("payment_method") or "",
                o.get("shipping_company") or "",
                float(o.get("total_amount") or 0),
                o.get("order_status") or "",
                CATEGORY_LABEL_AR.get(o.get("category"), o.get("category") or ""),
                o.get("city") or "",
                o.get("notes") or "",
            ])

        # Column widths
        widths = [16, 14, 26, 18, 22, 18, 14, 22, 12, 16, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

        # Summary sheet
        ws2 = wb.create_sheet("ملخّص")
        ws2.sheet_view.rightToLeft = True
        ws2.append(["الإجمالي", len(rows), "طلب"])
        ws2.append(["المجموع المالي", round(sum(float(o.get("total_amount") or 0) for o in rows), 2), "ر.س"])
        ws2.append([])
        ws2.append(["الفئة", "عدد الطلبات", "المجموع"])
        cat_agg: dict[str, dict] = {}
        for o in rows:
            c = o.get("category") or "_"
            b = cat_agg.setdefault(c, {"count": 0, "amount": 0.0})
            b["count"] += 1
            b["amount"] += float(o.get("total_amount") or 0)
        for c, b in cat_agg.items():
            ws2.append([CATEGORY_LABEL_AR.get(c, c), b["count"], round(b["amount"], 2)])
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 18

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
        filename = f"mezan_orders_{ts}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )

    @router.post("/{order_number}/resync")
    async def resync_order(order_number: str, user: dict = Depends(current_user)):
        """Re-fetch one order from Salla without invoking Qoyod writes."""
        from salla_integration.sync import resync_single_order
        from salla_integration.service import SallaError
        try:
            result = await resync_single_order(db, user["id"], order_number)
            if result.get("after"):
                overrides = await get_policy_map(db, user["id"])
                result["after"]["category"] = resolve_category(
                    result["after"].get("order_status"), overrides
                )
            return result
        except SallaError as e:
            raise HTTPException(
                status_code=e.status_code if e.status_code != 200 else 400,
                detail={"message": str(e), "needs_reauth": e.needs_reauth},
            )
        except Exception:
            error_reference = uuid.uuid4().hex
            logger.exception(
                "Unexpected single-order Salla resync failure "
                "error_reference=%s user_id=%s order_number=%s",
                error_reference,
                user.get("id"),
                order_number,
            )
            raise HTTPException(
                status_code=500,
                detail={
                    "message": "تعذر إعادة فحص الطلب من سلة",
                    "error_reference": error_reference,
                    "order_number": order_number,
                },
            )

    # Iter-91 Phase 2 — order adjustments audit log
    adj_router = APIRouter(prefix="/order-adjustments", tags=["orders"])

    @adj_router.get("")
    async def list_order_adjustments(
        order_number: Optional[str] = Query(None),
        reason: Optional[str] = Query(None),
        from_date: Optional[str] = Query(None),
        to_date: Optional[str] = Query(None),
        page: int = Query(1, ge=1),
        limit: int = Query(100, ge=1, le=500),
        user: dict = Depends(current_user),
    ):
        """Paginated list of recorded order modifications detected during
        resync (Iter-91 Phase 2). Each row captures total/COGS deltas
        plus the items added/removed/modified."""
        q: dict = {"user_id": user["id"]}
        if order_number:
            q["order_number"] = str(order_number).strip()
        if reason:
            q["reason"] = reason
        if from_date or to_date:
            d: dict = {}
            if from_date:
                d["$gte"] = from_date
            if to_date:
                d["$lte"] = to_date + "T23:59:59"
            q["created_at"] = d
        total = await db.order_adjustments.count_documents(q)
        skip = (page - 1) * limit
        items = await (
            db.order_adjustments
            .find(q, {"_id": 0})
            .sort([("created_at", -1)])
            .skip(skip).limit(limit)
            .to_list(limit)
        )
        return {
            "items": items, "total": total, "page": page, "limit": limit,
            "pages": max(1, (total + limit - 1) // limit),
        }

    parent_router.include_router(adj_router)

    parent_router.include_router(router)
