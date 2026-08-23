"""Excel parser for Salla platform exports.

The parser is forgiving: it auto-detects the relevant columns by checking common
Arabic and English header names that Salla uses in its exported sheets.
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional

import openpyxl

from payment_methods import (
    CARD_FALLBACK_SUB_KEYS,
    KNOWN_PAYMENT_SUB_KEYS,
    SALLA_SUB_KEYS,
    normalize_payment_method,
)


MAX_SALLA_ROWS = 50_000
MAX_SALLA_COLUMNS = 128


# Candidate column-name patterns (case-insensitive substring match)
TOTAL_COLS = [
    "إجمالي الطلب", "اجمالي الطلب", "إجمالي السلة", "اجمالي السلة",
    "المبلغ الإجمالي", "المبلغ الاجمالي", "الإجمالي", "الاجمالي",
    "total", "grand total", "order total", "amount",
]
PAYMENT_COLS = [
    "طريقة الدفع", "وسيلة الدفع", "بوابة الدفع",
    "payment method", "payment", "gateway",
]
SHIPPING_COLS = [
    "شركة الشحن", "وسيلة الشحن", "موفر الشحن", "مزود الشحن",
    "shipping company", "shipping", "courier", "carrier",
]
SOURCE_COLS = [
    "مصدر الطلب", "المصدر", "مصدر", "قناة البيع", "القناة",
    "مصدر الزيارة", "مصدر العميل", "نوع المصدر",
    "order source", "source", "channel", "platform", "referrer",
]
CUSTOMER_NAME_COLS = ["اسم العميل", "العميل", "الاسم", "اسم المشتري", "customer", "customer name", "buyer", "buyer name", "name"]
CUSTOMER_MOBILE_COLS = ["جوال العميل", "رقم الجوال", "الجوال", "الهاتف", "رقم الهاتف", "phone", "mobile", "customer phone", "customer mobile"]
SUBTOTAL_COLS = ["المجموع الفرعي", "السعر قبل الضريبة", "subtotal", "sub total", "items total"]
SHIPPING_COST_COLS = ["تكلفة الشحن", "رسوم الشحن", "shipping cost", "shipping fees", "shipping price"]
DISCOUNT_COLS = ["الخصم", "قيمة الخصم", "discount", "coupon"]
CURRENCY_COLS = ["العملة", "currency"]


def _normalize_currency(v) -> str:
    """Salla Excel exports sometimes write the currency as the integer
    code "1" or as a literal string "None" or as empty. The merchant
    sells exclusively in SAR, so we coerce any non-recognized value to
    "SAR" instead of letting garbage values pollute the database."""
    if v is None:
        return "SAR"
    s = str(v).strip()
    if not s or s in ("1", "None", "none", "null"):
        return "SAR"
    u = s.upper()
    if u in ("SAR", "USD", "EUR", "AED", "EGP", "KWD", "BHD", "QAR", "OMR"):
        return u
    return "SAR"

# Salla exports place the order source at column BA (index 52, 0-indexed) by default.
SALLA_SOURCE_COL_INDEX = 52  # Excel column "BA"
ORDER_ID_COLS = ["رقم الطلب", "order id", "order number", "id", "#"]
STATUS_COLS = [
    "حالة الطلب", "الحالة", "حالة طلب", "حاله الطلب", "حاله",
    "وضع الطلب", "وضع",
    "status", "order status", "order_status", "state",
]
DATE_COLS = [
    "تاريخ إنشاء الطلب", "تاريخ إنشاء", "تاريخ انشاء الطلب", "تاريخ انشاء",
    "تاريخ الطلب", "تاريخ الإنشاء", "تاريخ الانشاء",
    "التاريخ", "تاريخ",
    "date", "created at", "created_at", "order date", "order_date",
    "creation date", "order created at",
]
# Salla's standard Excel layout places the order creation date at column Q (index 16).
SALLA_DATE_COL_INDEX = 16  # Excel column "Q"


def _norm(s: object) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _find_header_row(rows: list[list]) -> int:
    """Return index of the row that looks like the header. Looks at first 10 rows."""
    best_idx = 0
    best_hits = -1
    keywords = TOTAL_COLS + PAYMENT_COLS + SHIPPING_COLS + ORDER_ID_COLS
    keywords_norm = [_norm(k) for k in keywords]
    for i, row in enumerate(rows[:15]):
        hits = 0
        for cell in row:
            v = _norm(cell)
            if not v:
                continue
            for kw in keywords_norm:
                if kw and kw in v:
                    hits += 1
                    break
        if hits > best_hits:
            best_hits = hits
            best_idx = i
    return best_idx


def _match_col(headers_norm: list[str], candidates: list[str]) -> Optional[int]:
    cand_norm = [_norm(c) for c in candidates]
    # exact-ish first (require non-empty header to avoid matching blank columns)
    for i, h in enumerate(headers_norm):
        if not h:
            continue
        for c in cand_norm:
            if c and c == h:
                return i
    # substring (require both sides non-empty so blank headers don't false-match)
    for i, h in enumerate(headers_norm):
        if not h:
            continue
        for c in cand_norm:
            if c and (c in h or h in c):
                return i
    return None


def _to_float(v: object) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v)
    # strip currency words and any non numeric except dot and minus
    s = re.sub(r"[^\d\.\-]", "", s.replace(",", "."))
    if not s or s in (".", "-"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_salla_excel(file_bytes: bytes) -> dict:
    """Parse a Salla-style Excel export and return aggregated raw data.

    Returns:
        {
          "total_sales": float,
          "total_orders": int,
          "payment_methods": [{"name": str, "orders_count": int, "total_sales": float}, ...],
          "shipping_companies": [{"name": str, "orders_count": int}, ...],
          "orders_sample": [...],
          "detected_columns": {...}
        }
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        data_only=True,
        read_only=True,
        keep_links=False,
    )
    try:
        ws = wb.active
        if ws.max_row > MAX_SALLA_ROWS + 15:
            raise ValueError("يتجاوز الملف الحد المسموح: 50,000 طلب")
        if ws.max_column > MAX_SALLA_COLUMNS:
            raise ValueError("يتجاوز الملف الحد المسموح: 128 عمود")
        rows = [
            list(r)
            for r in ws.iter_rows(
                values_only=True,
                max_row=MAX_SALLA_ROWS + 15,
                max_col=MAX_SALLA_COLUMNS,
            )
        ]
    finally:
        wb.close()
    if not rows:
        raise ValueError("الملف فارغ")

    header_idx = _find_header_row(rows)
    headers = rows[header_idx]
    headers_norm = [_norm(h) for h in headers]

    col_total = _match_col(headers_norm, TOTAL_COLS)
    col_payment = _match_col(headers_norm, PAYMENT_COLS)
    col_shipping = _match_col(headers_norm, SHIPPING_COLS)
    col_source = _match_col(headers_norm, SOURCE_COLS)
    col_order = _match_col(headers_norm, ORDER_ID_COLS)
    col_status = _match_col(headers_norm, STATUS_COLS)
    col_date = _match_col(headers_norm, DATE_COLS)
    col_customer = _match_col(headers_norm, CUSTOMER_NAME_COLS)
    col_mobile = _match_col(headers_norm, CUSTOMER_MOBILE_COLS)
    col_subtotal = _match_col(headers_norm, SUBTOTAL_COLS)
    col_ship_cost = _match_col(headers_norm, SHIPPING_COST_COLS)
    col_discount = _match_col(headers_norm, DISCOUNT_COLS)
    col_currency = _match_col(headers_norm, CURRENCY_COLS)

    # Salla's standard Excel layout: column A = order number, column B = order status.
    # If no header keyword matched, fall back to column index B (1).
    if col_status is None and len(headers_norm) > 1:
        col_status = 1

    # Fallback: Salla exports place order source at column BA (index 52)
    if col_source is None and len(headers) > SALLA_SOURCE_COL_INDEX:
        col_source = SALLA_SOURCE_COL_INDEX

    # Fallback: Salla exports place order creation date at column Q (index 16)
    if col_date is None and len(headers) > SALLA_DATE_COL_INDEX:
        col_date = SALLA_DATE_COL_INDEX

    if col_total is None:
        raise ValueError(
            "لم نتمكن من العثور على عمود إجمالي المبلغ في الملف. تأكد من أنه ملف طلبات سلة."
        )

    data_rows = rows[header_idx + 1 :]

    total_sales = 0.0
    total_orders = 0
    payments: dict[str, dict] = {}
    shippings: dict[str, dict] = {}
    sources: dict[str, dict] = {}
    sample_orders: list[dict] = []
    individual_orders: list[dict] = []  # NEW: every order row, used by unified-orders pipeline

    def _cell(col, row):
        if col is None or col >= len(row) or row[col] in (None, ""):
            return None
        return row[col]

    for row in data_rows:
        # skip totally empty rows
        if not any(cell not in (None, "") for cell in row):
            continue
        amount = _to_float(row[col_total] if col_total < len(row) else 0)
        if amount <= 0:
            # try detecting a row as valid even if amount is 0 only if there is order id
            if col_order is not None and col_order < len(row) and row[col_order]:
                pass
            else:
                continue

        payment_name = (
            str(row[col_payment]).strip()
            if col_payment is not None and col_payment < len(row) and row[col_payment]
            else "غير محدد"
        )
        # iter-72: scrub apostrophes / zero-width chars so DB stays clean.
        from shipping_companies import scrub_shipping_company
        shipping_name = (
            scrub_shipping_company(str(row[col_shipping]))
            if col_shipping is not None and col_shipping < len(row) and row[col_shipping]
            else "غير محدد"
        )
        source_name = (
            str(row[col_source]).strip()
            if col_source is not None and col_source < len(row) and row[col_source]
            else "غير محدد"
        )

        total_sales += amount
        total_orders += 1

        p = payments.setdefault(payment_name, {"name": payment_name, "orders_count": 0, "total_sales": 0.0})
        p["orders_count"] += 1
        p["total_sales"] += amount

        s = shippings.setdefault(shipping_name, {"name": shipping_name, "orders_count": 0})
        s["orders_count"] += 1

        src = sources.setdefault(source_name, {"name": source_name, "orders_count": 0, "total_sales": 0.0})
        src["orders_count"] += 1
        src["total_sales"] += amount

        order_id_val = str(_cell(col_order, row) or "").strip()
        status_val = str(_cell(col_status, row) or "").strip()
        date_raw = _cell(col_date, row)
        # openpyxl returns datetime/date objects when cell is formatted as Excel date.
        # Preserve the ISO date so downstream `_normalize_date_str` parses it correctly.
        if isinstance(date_raw, datetime):
            date_val = date_raw.date().isoformat()
        elif isinstance(date_raw, date):
            date_val = date_raw.isoformat()
        else:
            date_val = str(date_raw).strip() if date_raw is not None else ""

        if len(sample_orders) < 10:
            sample_orders.append({
                "order_id": order_id_val,
                "amount": amount,
                "payment_method": payment_name,
                "shipping_company": shipping_name,
                "status": status_val,
                "date": date_val,
            })

        # NEW: store full per-order data for the unified orders pipeline.
        # order_number is required as the dedup key — fall back to a synthesized
        # one if not present so we still preserve aggregates.
        if order_id_val:
            individual_orders.append({
                "order_number": order_id_val,
                "order_id": order_id_val,
                "order_date_raw": date_val,
                "order_status": status_val,
                "customer_name": str(_cell(col_customer, row) or "").strip(),
                "customer_mobile": str(_cell(col_mobile, row) or "").strip(),
                "payment_method": payment_name if payment_name != "غير محدد" else "",
                "shipping_company": shipping_name if shipping_name != "غير محدد" else "",
                "shipping_cost": _to_float(_cell(col_ship_cost, row) or 0),
                "subtotal": _to_float(_cell(col_subtotal, row) or 0),
                "discount": _to_float(_cell(col_discount, row) or 0),
                "total_amount": amount,
                "currency": _normalize_currency(_cell(col_currency, row)),
                "source": source_name if source_name != "غير محدد" else "",
            })

    if total_orders == 0:
        raise ValueError("لم نعثر على أي طلبات صالحة في الملف.")

    return {
        "total_sales": round(total_sales, 2),
        "total_orders": total_orders,
        "payment_methods": [
            {**v, "total_sales": round(v["total_sales"], 2)}
            for v in sorted(payments.values(), key=lambda x: -x["total_sales"])
        ],
        "shipping_companies": [
            v for v in sorted(shippings.values(), key=lambda x: -x["orders_count"])
        ],
        "order_sources": [
            {**v, "total_sales": round(v["total_sales"], 2)}
            for v in sorted(sources.values(), key=lambda x: -x["orders_count"])
        ],
        "orders_sample": sample_orders,
        "orders_individual": individual_orders,
        "detected_columns": {
            "total": headers[col_total] if col_total is not None else None,
            "payment": headers[col_payment] if col_payment is not None else None,
            "shipping": headers[col_shipping] if col_shipping is not None else None,
            "source": headers[col_source] if col_source is not None else None,
            "order_id": headers[col_order] if col_order is not None else None,
        },
    }


def normalize_name(name: str) -> str:
    """Lower-cased and stripped name for fuzzy matching settings vs file values."""
    if name is None:
        return ""
    n = str(name).strip().lower()
    n = re.sub(r"\s+", " ", n)
    # Remove Arabic diacritics
    n = re.sub(r"[\u064B-\u0652]", "", n)
    # Iteration 30: also unify common Arabic letter variants so that
    # "البطاقة الإئتمانية" matches "بطاقة ائتمانية" or "credit card".
    n = n.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    n = n.replace("ى", "ي").replace("ة", "ه").replace("ؤ", "و").replace("ئ", "ي")
    return n


# Iteration 30: payment-method synonym groups.
# Maps a CANONICAL token (Arabic preferred) to every alternate spelling
# Salla / the merchant might use across configs and order rows. Each
# group is bidirectional — if EITHER side appears in the settings name
# AND the other side appears in the order's payment_method, we match.
# Raw tokens here may contain ANY spelling — they're normalised at
# import time below so the matcher only ever compares like-with-like.
_RAW_PAYMENT_SYNONYMS: list[set[str]] = [
    # Mada
    {"مدى", "مدا", "مادا", "mada"},
    # Tamara
    {"تمارا", "تمارة", "tamara"},
    # Tabby
    {"تابي", "tabby"},
    # Emkan / Amkan
    {"امكان", "إمكان", "emkan", "amkan", "emkaninstallment", "emkan installment",
     "امكان للتقسيط", "إمكان للتقسيط"},
    # Apple Pay
    {"ابل باي", "ابل بي", "applepay", "apple pay", "apple"},
    # Google Pay
    {"جوجل باي", "قوقل باي", "googlepay", "google pay", "google"},
    # STC Pay
    {"اس تي سي باي", "stcpay", "stc pay", "stc"},
    # Credit / debit cards (Visa / MasterCard / generic credit)
    {"بطاقه ائتمانيه", "البطاقه الائتمانيه", "بطاقة ائتمانية",
     "البطاقة الإئتمانية", "credit card", "creditcard", "visa", "mastercard",
     "master card", "visa/mastercard", "visa / mastercard",
     "بطاقة بنكية", "بطاقه بنكيه"},
    # Cash on delivery
    {"الدفع عند الاستلام", "عند الاستلام", "عند الاستلم", "نقدا عند الاستلام",
     "cod", "cash on delivery", "cash_on_delivery", "cashondelivery"},
    # Bank transfer
    {"تحويل بنكي", "حواله بنكيه", "حوالة بنكية", "bank transfer", "banktransfer"},
    # Wallet (mostly Salla-specific)
    {"محفظه", "محفظة", "wallet", "salla wallet"},
]
# Pre-normalised groups — each token is run through normalize_name once.
PAYMENT_SYNONYMS: list[set[str]] = [
    {normalize_name(t) for t in group if t} for group in _RAW_PAYMENT_SYNONYMS
]


def _payment_synonym_match(settings_key: str, order_key: str) -> bool:
    """Iteration 30: bidirectional synonym lookup. Both keys are already
    passed through normalize_name(). Returns True if they belong to the
    SAME synonym group OR if one contains the other as substring (the
    legacy fuzzy fallback)."""
    if not settings_key or not order_key:
        return False
    if settings_key == order_key:
        return True
    # Substring fallback (legacy behaviour, kept for backwards-compat).
    if settings_key in order_key or order_key in settings_key:
        return True
    # Synonym groups: pick the group of the settings key, see if the
    # order key resolves into the same group. Both keys are already
    # normalised so we can do plain string membership / substring checks.
    for group in PAYMENT_SYNONYMS:
        in_settings = any(
            t == settings_key or t in settings_key or settings_key in t
            for t in group
        )
        if not in_settings:
            continue
        in_order = any(
            t == order_key or t in order_key or order_key in t
            for t in group
        )
        if in_order:
            return True
    return False


def match_settings(
    parsed: dict,
    payment_settings: list[dict],
    shipping_settings: list[dict],
) -> dict:
    """Cross-reference parsed totals with user-provided commission rates and shipping costs.

    Salla payment fee formula (when individual orders are available):
        unrounded_order_fee = order_amount * commission_percent / 100 + fixed_fee
        base_commission     = sum(round(unrounded_order_fee, 2))
        vat_amount          = sum(round(unrounded_order_fee * vat_percent / 100, 2))
        fee_amount          = base_commission + vat_amount

    Other providers and aggregate-only callers retain the historical aggregate
    formula.  Per-order rounding is the authoritative Salla invoice method.

    Unmatched names get 0 / 0 / 0 but are still returned.
    """
    # Build lookups by exact display spelling and canonical sub-method.  Exact
    # canonical matching prevents a generic card row from accidentally taking
    # Visa's fee merely because the legacy synonym group contains both.
    payment_map: dict[str, dict] = {}
    canonical_payment_map: dict[str, dict] = {}
    for p in payment_settings:
        config = {
            "commission_percent": float(p.get("commission_percent", 0) or 0),
            "fixed_fee": float(p.get("fixed_fee", 0) or 0),
            "vat_percent": float(p.get("vat_percent", 0) or 0),
        }
        payment_map[normalize_name(p["name"])] = config
        sub_key, _display, _parent = normalize_payment_method(p.get("name") or "")
        if sub_key in KNOWN_PAYMENT_SUB_KEYS:
            canonical_payment_map.setdefault(sub_key, config)
    shipping_map = {
        normalize_name(s["name"]): {
            "cost_per_order": float(s.get("cost_per_order", 0) or 0),
            "vat_percent": float(s.get("vat_percent", 0) or 0),
            "is_deferred": bool(s.get("is_deferred", False)),
        }
        for s in shipping_settings
    }

    payment_breakdown = []
    total_payment_fees = 0.0
    individual_orders = parsed.get("orders_individual") or []

    def _same_payment_method(raw: str, grouped_name: str) -> bool:
        raw_key = normalize_name(raw)
        grouped_key = normalize_name(grouped_name)
        if raw_key == grouped_key:
            return True
        raw_sub, _raw_display, _raw_parent = normalize_payment_method(raw)
        grouped_sub, _grouped_display, _grouped_parent = normalize_payment_method(grouped_name)
        if raw_sub in KNOWN_PAYMENT_SUB_KEYS and grouped_sub in KNOWN_PAYMENT_SUB_KEYS:
            return raw_sub == grouped_sub
        return _payment_synonym_match(raw_key, grouped_key)

    cent = Decimal("0.01")

    def _round_money(value: Decimal) -> Decimal:
        return value.quantize(cent, rounding=ROUND_HALF_UP)

    for pm in parsed["payment_methods"]:
        key = normalize_name(pm["name"])
        cfg = payment_map.get(key)
        order_sub_key, _display, _parent = normalize_payment_method(pm["name"])
        if cfg is None and order_sub_key in KNOWN_PAYMENT_SUB_KEYS:
            cfg = canonical_payment_map.get(order_sub_key)
            if cfg is None and order_sub_key in CARD_FALLBACK_SUB_KEYS:
                cfg = canonical_payment_map.get("credit_card")
        # Iteration 30 fallback: use broad synonym/fuzzy matching only for an
        # unknown legacy label.  Two different known rails stay distinct.
        if cfg is None and order_sub_key not in KNOWN_PAYMENT_SUB_KEYS:
            for k, v in payment_map.items():
                if _payment_synonym_match(k, key):
                    cfg = v
                    break
        matched = cfg is not None
        cfg = cfg or {"commission_percent": 0.0, "fixed_fee": 0.0, "vat_percent": 0.0}

        pct = cfg["commission_percent"]
        fixed = cfg["fixed_fee"]
        vat_pct = cfg["vat_percent"]

        method_orders = [
            order for order in individual_orders
            if _same_payment_method(
                str(order.get("payment_method") or ""),
                str(pm.get("name") or ""),
            )
        ]
        if (
            order_sub_key in SALLA_SUB_KEYS
            and method_orders
            and len(method_orders) == int(pm.get("orders_count") or 0)
        ):
            # Salla invoice semantics: calculate on every positive sale,
            # round the fee per order, and calculate VAT from the unrounded
            # fee before independently rounding VAT.  Negative refund rows in
            # the supplied invoices carry no new fee or VAT.
            pct_decimal = Decimal(str(pct)) / Decimal("100")
            fixed_decimal = Decimal(str(fixed))
            vat_decimal = Decimal(str(vat_pct)) / Decimal("100")
            base_total = Decimal("0")
            vat_total = Decimal("0")
            for order in method_orders:
                raw_amount = order.get("total_amount")
                if raw_amount is None:
                    raw_amount = order.get("amount")
                amount = Decimal(str(raw_amount or 0))
                if amount <= 0:
                    continue
                unrounded_fee = amount * pct_decimal + fixed_decimal
                base_total += _round_money(unrounded_fee)
                vat_total += _round_money(unrounded_fee * vat_decimal)
            base_commission = float(_round_money(base_total))
            vat_amount = float(_round_money(vat_total))
            calculation_basis = "per_order_salla_rounding"
        else:
            # Backwards-compatible fallback for callers that only provide
            # method aggregates and no individual order rows.
            base_commission = round(
                pm["total_sales"] * pct / 100.0 + pm["orders_count"] * fixed,
                2,
            )
            vat_amount = round(base_commission * vat_pct / 100.0, 2)
            calculation_basis = "aggregate_fallback"
        fee_amount = round(base_commission + vat_amount, 2)
        total_payment_fees += fee_amount

        payment_breakdown.append({
            "name": pm["name"],
            "orders_count": pm["orders_count"],
            "total_sales": pm["total_sales"],
            "commission_percent": pct,
            "fixed_fee": fixed,
            "vat_percent": vat_pct,
            "base_commission": base_commission,
            "vat_amount": vat_amount,
            "fee_amount": fee_amount,
            "net_amount": round(pm["total_sales"] - fee_amount, 2),
            "matched": matched,
            "fee_calculation_basis": calculation_basis,
        })

    shipping_breakdown = []
    total_shipping_cost = 0.0
    deferred_shipping_cost = 0.0
    for sc in parsed["shipping_companies"]:
        key = normalize_name(sc["name"])
        cfg = shipping_map.get(key)
        if cfg is None:
            for k, v in shipping_map.items():
                if k and (k in key or key in k):
                    cfg = v
                    break
        matched = cfg is not None
        cfg = cfg or {"cost_per_order": 0.0, "vat_percent": 0.0, "is_deferred": False}

        cost = cfg["cost_per_order"]
        vat_pct = cfg["vat_percent"]
        is_deferred = bool(cfg.get("is_deferred", False))
        base_cost = round(cost * sc["orders_count"], 2)
        vat_amount = round(base_cost * vat_pct / 100.0, 2)
        total = round(base_cost + vat_amount, 2)
        total_shipping_cost += total
        if is_deferred:
            deferred_shipping_cost += total
        shipping_breakdown.append({
            "name": sc["name"],
            "orders_count": sc["orders_count"],
            "cost_per_order": cost,
            "base_cost": base_cost,
            "vat_percent": vat_pct,
            "vat_amount": vat_amount,
            "total_cost": total,
            "matched": matched,
            "is_deferred": is_deferred,
        })

    return {
        "payment_breakdown": payment_breakdown,
        "shipping_breakdown": shipping_breakdown,
        "total_payment_fees": round(total_payment_fees, 2),
        "total_shipping_cost": round(total_shipping_cost, 2),
        "deferred_shipping_cost": round(deferred_shipping_cost, 2),
    }
