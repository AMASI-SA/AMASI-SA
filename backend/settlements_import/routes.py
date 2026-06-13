"""HTTP endpoints for /api/payment-settlements/* (Phase 80)."""
from __future__ import annotations

from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile

from auth import get_current_user_from_db

from .service import (
    coverage_analytics,
    delete_file,
    ensure_settlements_indexes,
    get_file_detail,
    import_file,
    list_files,
)


MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB — settlement files are tiny xlsx


def attach_payment_settlements_routes(api_router: APIRouter, db) -> None:
    router = APIRouter(prefix="/payment-settlements", tags=["payment-settlements"])

    async def current_user(request: Request) -> dict:
        return await get_current_user_from_db(request, db)

    # ── Upload ─────────────────────────────────────────────────────────
    @router.post("/upload")
    async def upload(
        file: UploadFile = File(...),
        provider_hint: Optional[str] = Form(default=None),
        invoice_date: Optional[str] = Form(default=None),
        user: dict = Depends(current_user),
    ):
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="الملف فارغ.")
        if len(content) > MAX_FILE_BYTES:
            raise HTTPException(status_code=413, detail="حجم الملف يتجاوز 10 ميجابايت.")
        try:
            result = await import_file(
                db, user["id"],
                filename=file.filename or "settlement.xlsx",
                content=content,
                provider_hint=(provider_hint or None),
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        # Iter-159g — Persist the merchant-supplied invoice issue date as
        # `header.settlement_date` so the unified overview's
        # "تاريخ التحويل" column shows the platform issue date instead of
        # the upload timestamp.
        fid = result.get("file_id") if isinstance(result, dict) else None
        if fid and invoice_date:
            import re as _re
            iso = str(invoice_date).strip()[:10]
            if _re.match(r"^\d{4}-\d{2}-\d{2}$", iso):
                await db.settlement_files.update_one(
                    {"id": fid, "user_id": user["id"]},
                    {"$set": {"header.settlement_date": iso}},
                )
                if isinstance(result, dict):
                    result["invoice_date"] = iso
        return result

    # ── List uploaded files ────────────────────────────────────────────
    @router.get("")
    async def list_settlement_files(
        limit: int = 50,
        user: dict = Depends(current_user),
    ):
        files = await list_files(db, user["id"], limit=limit)
        return {"files": files}

    # ── File detail (incl. unmatched orders) ──────────────────────────
    @router.get("/{file_id}")
    async def file_detail(file_id: str, user: dict = Depends(current_user)):
        doc = await get_file_detail(db, user["id"], file_id)
        if not doc:
            raise HTTPException(status_code=404, detail="الملف غير موجود.")
        return doc

    # ── Delete + rollback actual_* on orders ──────────────────────────
    @router.delete("/{file_id}")
    async def remove_file(file_id: str, user: dict = Depends(current_user)):
        result = await delete_file(db, user["id"], file_id)
        if not result.get("removed"):
            raise HTTPException(status_code=404, detail="الملف غير موجود.")
        return {"ok": True, **result}

    # ── Iter-159e — Manual override of settlement/transfer date ───────
    # For providers whose files don't contain a transfer date (e.g.
    # Salla payment invoices), the merchant can set/edit it manually so
    # the unified overview correctly shows when the money actually
    # landed in the bank instead of the file upload timestamp.
    @router.patch("/{file_id}/settlement-date")
    async def patch_settlement_date(
        file_id: str,
        payload: dict,
        user: dict = Depends(current_user),
    ):
        new_date = (payload or {}).get("settlement_date")
        if new_date in (None, ""):
            # Allow clearing the manual override → fall back to derived.
            new_date = None
        else:
            new_date = str(new_date).strip()[:10]
            # Lightweight validation: YYYY-MM-DD shape only.
            import re as _re
            if not _re.match(r"^\d{4}-\d{2}-\d{2}$", new_date):
                raise HTTPException(
                    status_code=400,
                    detail="صيغة التاريخ يجب أن تكون YYYY-MM-DD")
        res = await db.settlement_files.update_one(
            {"id": file_id, "user_id": user["id"]},
            {"$set": {"header.settlement_date": new_date}},
        )
        if res.matched_count == 0:
            raise HTTPException(status_code=404, detail="الملف غير موجود.")
        return {"ok": True, "settlement_date": new_date}

    # ── Coverage analytics (estimated vs actual) ──────────────────────
    @router.get("/_analytics/coverage")
    async def analytics(user: dict = Depends(current_user)):
        return await coverage_analytics(db, user["id"])

    # ── Iter-156 — Salla-specific analytics for the new
    # SallaSettlements page (file list + per-payment-method breakdown).
    @router.get("/_analytics/salla")
    async def salla_analytics(
        user: dict = Depends(current_user),
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
    ):
        files = await db.settlement_files.find(
            {"user_id": user["id"], "provider": "salla"},
            {"_id": 0},
        ).sort("uploaded_at", -1).to_list(500)

        # Aggregate per-payment-method across all entries belonging to
        # Salla files of this user.
        entries_pipeline = [
            {"$match": {"user_id": user["id"], "provider": "salla",
                        "event_type": {"$in": ["sale", "refund"]}}},
            {"$group": {
                "_id": "$actual_payment_method",
                "count": {"$sum": 1},
                "gross": {"$sum": "$actual_gross_amount"},
                "fees": {"$sum": "$actual_payment_fee"},
                "vat": {"$sum": "$actual_payment_vat"},
                "net": {"$sum": "$actual_net_amount"},
                "refund_full": {"$sum": "$actual_refund_amount"},
                "refund_partial": {"$sum": "$actual_partial_refund_amount"},
            }},
            {"$sort": {"net": -1}},
        ]
        per_method = []
        async for r in db.settlement_entries.aggregate(entries_pipeline):
            per_method.append({
                "payment_method": r["_id"] or "unknown",
                "count": r["count"],
                "gross": round(r.get("gross") or 0, 2),
                "fees": round(r.get("fees") or 0, 2),
                "vat": round(r.get("vat") or 0, 2),
                "net": round(r.get("net") or 0, 2),
                "refund_full": round(r.get("refund_full") or 0, 2),
                "refund_partial": round(r.get("refund_partial") or 0, 2),
                "effective_fee_rate": round(
                    (r["fees"] / r["gross"]) * 100 if (r.get("gross") or 0) > 0 else 0,
                    2,
                ),
            })

        totals = {
            "files": len(files),
            "gross": round(sum(m["gross"] for m in per_method), 2),
            "fees": round(sum(m["fees"] for m in per_method), 2),
            "vat": round(sum(m["vat"] for m in per_method), 2),
            "net": round(sum(m["net"] for m in per_method), 2),
            "refund_full": round(sum(m["refund_full"] for m in per_method), 2),
            "refund_partial": round(sum(m["refund_partial"] for m in per_method), 2),
        }
        return {
            "files": files,
            "per_method": per_method,
            "totals": totals,
        }

    # ── Iter-158 — Unified weekly settlements view (Salla + Tamara + Tabby)
    # with month navigation + Excel export of selected entries.
    @router.get("/_overview/unified")
    async def unified_settlements_list(
        year: int,
        month: int,
        user: dict = Depends(current_user),
    ):
        import calendar
        if not (1 <= month <= 12):
            raise HTTPException(400, "month must be 1-12")
        last_day = calendar.monthrange(year, month)[1]
        from_dt = f"{year:04d}-{month:02d}-01"
        to_dt = f"{year:04d}-{month:02d}-{last_day:02d}"

        files = await db.settlement_files.find(
            {"user_id": user["id"]},
            {"_id": 0},
        ).sort("uploaded_at", -1).to_list(2000)

        # Iter-159e — Build {file_id → max entry settlement_date} so the
        # "تاريخ التحويل" column shows the actual settlement/transfer date
        # captured from the file rows (e.g. Tabby `Transfer Date`,
        # Tamara `event_date`) instead of the file upload timestamp.
        file_ids = [f["id"] for f in files]
        entry_settlement_dates: dict[str, str] = {}
        if file_ids:
            cur = db.settlement_entries.aggregate([
                {"$match": {"user_id": user["id"],
                            "file_id": {"$in": file_ids},
                            "settlement_date": {"$ne": None}}},
                {"$group": {"_id": "$file_id",
                            "max_date": {"$max": "$settlement_date"}}},
            ])
            async for r in cur:
                v = r.get("max_date")
                if v is None:
                    continue
                if hasattr(v, "isoformat"):
                    v = v.isoformat()
                entry_settlement_dates[r["_id"]] = str(v)[:10]

        rows = []
        for f in files:
            uploaded_at = f.get("uploaded_at")
            if hasattr(uploaded_at, "isoformat"):
                uploaded_at = uploaded_at.isoformat()
            uploaded_at = str(uploaded_at or "")

            manual = f.get("header", {}).get("settlement_date")
            derived = entry_settlement_dates.get(f["id"])
            transfer = f.get("header", {}).get("transfer_date")
            if manual:
                settlement_date = manual
                date_source = "manual"
            elif derived:
                settlement_date = derived
                date_source = "file_rows"
            elif transfer:
                settlement_date = transfer
                date_source = "file_rows"
            else:
                settlement_date = uploaded_at[:10]
                date_source = "uploaded_at"  # ⚠ requires manual override

            sd = (settlement_date or "")[:10]
            if not (from_dt <= sd <= to_dt):
                continue
            totals = f.get("totals", {}) or {}
            rows.append({
                "file_id": f["id"],
                "provider": f.get("provider"),
                "filename": f.get("filename"),
                "invoice_number": (f.get("header") or {}).get("invoice_number"),
                "payment_method": (f.get("header") or {}).get("payment_method", "متعدد"),
                "settlement_date": sd,
                "settlement_date_source": date_source,
                "gross": round(totals.get("gross") or 0, 2),
                "fees": round(totals.get("fees") or 0, 2),
                "vat": round(totals.get("vat") or 0, 2),
                "net_to_bank": round(totals.get("net") or 0, 2),
                "rows": f.get("rows", 0),
                "matched": f.get("matched", 0),
                "uploaded_at": uploaded_at,
            })

        rows.sort(key=lambda r: (r["settlement_date"] or "", r["uploaded_at"] or ""), reverse=True)

        provider_totals = {}
        for r in rows:
            p = r["provider"]
            t = provider_totals.setdefault(p, {"count": 0, "gross": 0, "fees": 0, "net": 0})
            t["count"] += 1
            t["gross"] += r["gross"]
            t["fees"] += r["fees"]
            t["net"] += r["net_to_bank"]
        for p in provider_totals:
            provider_totals[p] = {
                k: round(v, 2) if k != "count" else v
                for k, v in provider_totals[p].items()
            }

        return {
            "rows": rows,
            "year": year,
            "month": month,
            "provider_totals": provider_totals,
            "grand_total": {
                "count": len(rows),
                "gross": round(sum(r["gross"] for r in rows), 2),
                "fees": round(sum(r["fees"] for r in rows), 2),
                "net": round(sum(r["net_to_bank"] for r in rows), 2),
            },
        }

    @router.post("/_overview/export-excel")
    async def export_settlements_excel(
        payload: dict,
        user: dict = Depends(current_user),
    ):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from fastapi.responses import StreamingResponse

        ids = payload.get("file_ids") or []
        if not ids:
            raise HTTPException(400, "اختر تسوية واحدة على الأقل")
        files = await db.settlement_files.find(
            {"user_id": user["id"], "id": {"$in": ids}},
            {"_id": 0},
        ).to_list(2000)
        if not files:
            raise HTTPException(404, "لا توجد تسويات مطابقة")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "التسويات"
        ws.sheet_view.rightToLeft = True
        headers = [
            "المزوّد", "رقم الفاتورة", "تاريخ التسوية",
            "طريقة الدفع", "إجمالي المبيعات", "العمولات",
            "VAT", "صافي التحويل", "اسم الملف",
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="065F46")
            c.alignment = Alignment(horizontal="center")
        for f in files:
            totals = f.get("totals", {}) or {}
            uploaded_at = f.get("uploaded_at")
            if hasattr(uploaded_at, "isoformat"):
                uploaded_at = uploaded_at.isoformat()
            uploaded_at = str(uploaded_at or "")
            ws.append([
                (f.get("provider") or "").upper(),
                (f.get("header") or {}).get("invoice_number") or "—",
                ((f.get("header") or {}).get("settlement_date")
                 or uploaded_at[:10]),
                (f.get("header") or {}).get("payment_method") or "متعدد",
                round(totals.get("gross") or 0, 2),
                round(totals.get("fees") or 0, 2),
                round(totals.get("vat") or 0, 2),
                round(totals.get("net") or 0, 2),
                f.get("filename"),
            ])
        # Totals row
        ws.append([
            "الإجمالي", "", "", "",
            round(sum((f.get("totals") or {}).get("gross") or 0 for f in files), 2),
            round(sum((f.get("totals") or {}).get("fees") or 0 for f in files), 2),
            round(sum((f.get("totals") or {}).get("vat") or 0 for f in files), 2),
            round(sum((f.get("totals") or {}).get("net") or 0 for f in files), 2),
            "",
        ])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="D1FAE5")
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[col_letter].width = 18

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition":
                    'attachment; filename="settlements_export.xlsx"',
            },
        )

    api_router.include_router(router)
