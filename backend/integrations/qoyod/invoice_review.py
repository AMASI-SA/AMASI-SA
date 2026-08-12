"""Read-only Qoyod invoice review for Mezan.

This module deliberately has no send/approval surface.  It reads the local
``qoyod_invoices`` mirror, joins an invoice to Salla only when its Qoyod
``reference`` exactly equals the Salla ``order_number``, and can refresh the
mirror through the existing GET-only Qoyod synchroniser.

Tenant scopes are intentionally split:

* Qoyod invoices and credentials belong to the Qoyod marker tenant (currently
  ``main``).
* Salla orders belong to the authenticated merchant owner, resolved by
  :func:`integrations.qoyod.orders_owner.orders_owner_id` at the route layer.

No function in this module calls Qoyod POST/PUT/PATCH/DELETE.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from io import BytesIO
import math
import re
from typing import Any, Optional

try:
    from zoneinfo import ZoneInfo
    _RIYADH_TZ = ZoneInfo("Asia/Riyadh")
except (ImportError, KeyError):  # pragma: no cover
    _RIYADH_TZ = timezone(timedelta(hours=3))

from integrations.qoyod.eligible_orders import _normalize_status, _parse_iso_date
from integrations.qoyod.unsent_orders import _is_real


INVOICE_REVIEW_FLOOR = date(2026, 7, 1)
DEFAULT_PAGE_SIZE = 15
MAX_PAGE_SIZE = 100
MAX_EXPORT_ROWS = 50_000


def _eligible_salla_status(row: dict) -> bool:
    """The three explicitly-approved invoice-review states only."""
    for value in (row.get("order_status_slug"), row.get("order_status")):
        status = _normalize_status(value)
        if status in {
            "completed", "تم التنفيذ", "منتهي", "مكتمل",
            "delivered", "تم التوصيل",
            "in delivery", "shipping",
            "جاري التوصيل", "جارٍ التوصيل",
        }:
            return True
    return False


def parse_review_range(
    from_date: Optional[str],
    to_date: Optional[str],
    *,
    today: Optional[date] = None,
) -> tuple[date, date]:
    """Validate and clamp a review range to the accounting floor."""
    upper_default = today or datetime.now(_RIYADH_TZ).date()
    try:
        start = date.fromisoformat(from_date) if from_date \
            else INVOICE_REVIEW_FLOOR
        end = date.fromisoformat(to_date) if to_date else upper_default
    except (TypeError, ValueError) as exc:
        raise ValueError("date_must_be_iso_yyyy_mm_dd") from exc
    start = max(start, INVOICE_REVIEW_FLOOR)
    if end < INVOICE_REVIEW_FLOOR:
        raise ValueError("to_date_before_qoyod_invoice_review_floor")
    if start > end:
        raise ValueError("from_date_must_not_be_after_to_date")
    return start, end


def _invoice_issue_date(row: dict) -> Optional[date]:
    return _parse_iso_date(row.get("issue_date"))


def _invoice_in_range(row: dict, start: date, end: date) -> bool:
    issue_date = _invoice_issue_date(row)
    return issue_date is not None and start <= issue_date <= end


def _real_local_invoice(row: dict) -> bool:
    return _is_real(row.get("qoyod_invoice_id"))


def _reference(row: dict) -> str:
    """Use only Qoyod's canonical reference for an exact Salla join."""
    return str(row.get("reference") or "").strip()


def _search_matches(row: dict, search: Optional[str]) -> bool:
    needle = str(search or "").strip().casefold()
    if not needle:
        return True
    fields = (
        row.get("reference"),
        row.get("invoice_number"),
        row.get("qoyod_invoice_id"),
        row.get("customer_name"),
    )
    return any(needle in str(value or "").casefold() for value in fields)


async def _load_salla_orders(
    db,
    *,
    orders_user_id: str,
    start: date,
    end: date,
) -> tuple[dict[str, dict], int]:
    """Return every owner order for joins plus the eligible-range count.

    The exact-reference join is intentionally *not* restricted by Salla date
    or status: an in-range Qoyod invoice can legitimately point to an older or
    currently-ineligible Salla order and must still show that the order exists.
    Only the summary count is restricted to the selected creation-date window
    and the three approved states (completed / in_delivery / delivered).
    """
    query = {
        "user_id": orders_user_id,
    }
    projection = {
        "_id": 0,
        "order_number": 1,
        "order_status": 1,
        "order_status_slug": 1,
        "order_date": 1,
        "total_amount": 1,
        "customer_name": 1,
    }
    orders: dict[str, dict] = {}
    eligible_order_numbers: set[str] = set()
    cursor = db.unified_orders.find(query, projection).sort(
        "order_date", -1
    )
    async for row in cursor:
        order_number = str(row.get("order_number") or "").strip()
        # ``order_date`` is Salla's order creation date in unified_orders.
        # Never fall back to Mezan ``created_at`` for this report.
        created = _parse_iso_date(row.get("order_date"))
        if not order_number:
            continue
        orders.setdefault(order_number, row)
        if (
            created is not None
            and start <= created <= end
            and _eligible_salla_status(row)
        ):
            eligible_order_numbers.add(order_number)
    return orders, len(eligible_order_numbers)


async def _load_local_invoices(
    db,
    *,
    markers_user_id: str,
    start: date,
    end: date,
) -> list[dict]:
    """Load real invoices in range from the Qoyod tenant's local mirror."""
    query = {
        "user_id": markers_user_id,
    }
    projection = {
        "_id": 0,
        "qoyod_invoice_id": 1,
        "invoice_number": 1,
        "reference": 1,
        "customer_name": 1,
        "issue_date": 1,
        "due_date": 1,
        "currency": 1,
        "total": 1,
        "paid_amount": 1,
        "remaining": 1,
        "status": 1,
        "last_sync_at": 1,
        "raw_response.due_date": 1,
        "raw_response.currency": 1,
    }
    invoices: list[dict] = []
    cursor = db.qoyod_invoices.find(query, projection).sort(
        [("issue_date", -1), ("qoyod_invoice_id", -1)]
    )
    async for row in cursor:
        if not _real_local_invoice(row):
            continue
        if not _invoice_in_range(row, start, end):
            continue
        invoices.append(row)
    return invoices


def _safe_money(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _excel_safe_value(value: Any) -> Any:
    """Prevent formula injection without changing dates or numbers.

    Qoyod customer names, references and labels are external input.  Excel
    interprets text beginning with ``= + - @`` as a formula or command, so
    prefix such strings with an apostrophe before writing the workbook.
    Numeric/date objects are returned unchanged.
    """
    if not isinstance(value, str):
        return value
    candidate = value.lstrip()
    if candidate.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _invoice_item(row: dict, salla_order: Optional[dict]) -> dict:
    raw = row.get("raw_response") or {}
    reference = _reference(row)
    exact = bool(salla_order) and reference == str(
        salla_order.get("order_number") or ""
    ).strip()
    return {
        "qoyod_invoice_id": str(row.get("qoyod_invoice_id") or ""),
        "invoice_number": str(row.get("invoice_number") or "") or None,
        "reference": reference or None,
        "salla_order_number": (
            str(salla_order.get("order_number")) if exact else None
        ),
        "customer_name": row.get("customer_name"),
        "issue_date": _json_value(row.get("issue_date")),
        "due_date": _json_value(
            row.get("due_date") or raw.get("due_date")
        ),
        "currency": row.get("currency") or raw.get("currency") or None,
        "total": _safe_money(row.get("total")),
        "paid_amount": _safe_money(row.get("paid_amount")),
        "remaining": _safe_money(row.get("remaining")),
        "status": row.get("status"),
        "last_sync_at": _json_value(row.get("last_sync_at")),
        "exact_reference_match": exact,
        "salla_status": (
            salla_order.get("order_status_slug")
            or salla_order.get("order_status")
            if exact else None
        ),
        "salla_total": (
            _safe_money(salla_order.get("total_amount"))
            if exact else None
        ),
    }


async def build_invoice_review(
    db,
    *,
    orders_user_id: str,
    markers_user_id: str,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
    include_all: bool = False,
) -> dict:
    """Build the standalone invoice list and date-range summary."""
    start, end = parse_review_range(from_date, to_date)
    page = max(1, int(page))
    page_size = max(1, min(int(page_size), MAX_PAGE_SIZE))

    all_salla_orders, eligible_count = await _load_salla_orders(
        db,
        orders_user_id=orders_user_id,
        start=start,
        end=end,
    )
    invoices = await _load_local_invoices(
        db,
        markers_user_id=markers_user_id,
        start=start,
        end=end,
    )

    items: list[dict] = []
    exact_matches = 0
    for invoice in invoices:
        reference = _reference(invoice)
        salla_order = all_salla_orders.get(reference)
        item = _invoice_item(invoice, salla_order)
        if item["exact_reference_match"]:
            exact_matches += 1
        if _search_matches(invoice, search):
            items.append(item)

    total = len(items)
    if include_all:
        if total > MAX_EXPORT_ROWS:
            raise ValueError(
                f"invoice_export_exceeds_{MAX_EXPORT_ROWS}_rows"
            )
        visible = items
    else:
        offset = (page - 1) * page_size
        visible = items[offset:offset + page_size]

    sync_timestamps = [
        str(_json_value(invoice.get("last_sync_at")))
        for invoice in invoices if invoice.get("last_sync_at")
    ]
    latest_sync = max(sync_timestamps) if sync_timestamps else None
    return {
        "ok": True,
        "from_date": start.isoformat(),
        "to_date": end.isoformat(),
        "search": str(search or "").strip() or None,
        "page": page,
        "page_size": page_size,
        "total": total,
        "pages": max(1, math.ceil(total / page_size)),
        "summary": {
            "eligible_salla_orders": eligible_count,
            "qoyod_invoices": len(invoices),
            "exact_reference_matches": exact_matches,
            "unmatched_qoyod_invoices": len(invoices) - exact_matches,
            "latest_sync_at": latest_sync,
        },
        "last_sync_at": latest_sync,
        "items": visible,
        "sync_summary": {"ran": False},
    }


async def sync_invoice_review(
    db,
    *,
    orders_user_id: str,
    markers_user_id: str,
    api_client: Any,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
) -> dict:
    """Refresh through the existing GET-only sync, then return the review."""
    start, _ = parse_review_range(from_date, to_date)
    from integrations.qoyod.qoyod_invoices_sync import sync_qoyod_invoices

    sync_summary = await sync_qoyod_invoices(
        db,
        user_id=markers_user_id,
        api_client=api_client,
        from_date=start,
    )
    sync_summary["ran"] = True
    if not sync_summary.get("ok"):
        return {
            "ok": False,
            "error": "qoyod_invoice_read_sync_failed",
            "sync_summary": sync_summary,
            "items": [],
        }
    result = await build_invoice_review(
        db,
        orders_user_id=orders_user_id,
        markers_user_id=markers_user_id,
        from_date=from_date,
        to_date=to_date,
        search=search,
        page=page,
        page_size=page_size,
    )
    result["sync_summary"] = sync_summary
    return result


def build_invoice_review_workbook(report: dict) -> bytes:
    """Return a polished RTL Excel workbook for the selected filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    invoices = workbook.active
    invoices.title = "فواتير قيود"
    invoices.sheet_view.rightToLeft = True
    invoices.freeze_panes = "A3"
    invoices.auto_filter.ref = f"A2:P{max(2, len(report['items']) + 2)}"

    navy = "071128"
    blue = "125C8E"
    pale = "EAF4F8"
    green = "0A8F67"
    red = "C5364A"
    white = "FFFFFF"
    thin = Side(style="thin", color="D7E0E8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "رقم فاتورة قيود", "رقم الفاتورة",
        "مرجع طلب سلة", "اسم العميل",
        "تاريخ الإصدار", "تاريخ الاستحقاق", "العملة",
        "القيمة الإجمالية", "المدفوع",
        "الرصيد", "حالة الفاتورة",
        "مطابقة المرجع", "حالة طلب سلة",
        "إجمالي طلب سلة",
        "آخر مزامنة", "فرق الإجمالي",
    ]
    invoices.merge_cells(start_row=1, start_column=1,
                         end_row=1, end_column=len(headers))
    title = invoices.cell(1, 1, "تقرير فواتير قيود — ميزان")
    title.fill = PatternFill("solid", fgColor=navy)
    title.font = Font(name="Arial", color=white, bold=True, size=16)
    title.alignment = Alignment(horizontal="right", vertical="center")
    invoices.row_dimensions[1].height = 32

    for index, header in enumerate(headers, 1):
        cell = invoices.cell(2, index, header)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(name="Arial", color=white, bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    invoices.row_dimensions[2].height = 28

    for row_index, item in enumerate(report.get("items", []), 3):
        salla_total = item.get("salla_total")
        qoyod_total = item.get("total")
        difference = (
            round(float(salla_total) - float(qoyod_total), 2)
            if salla_total is not None and qoyod_total is not None else None
        )
        values = [
            item.get("qoyod_invoice_id"), item.get("invoice_number"),
            item.get("reference"), item.get("customer_name"),
            item.get("issue_date"), item.get("due_date"),
            item.get("currency"), item.get("total"),
            item.get("paid_amount"), item.get("remaining"),
            item.get("status"),
            "مطابق" if item.get("exact_reference_match") else "غير مطابق",
            item.get("salla_status"), item.get("salla_total"),
            item.get("last_sync_at"), difference,
        ]
        for col_index, value in enumerate(values, 1):
            cell = invoices.cell(
                row_index, col_index, _excel_safe_value(value)
            )
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = border
            if row_index % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F8FBFD")
        for col_index in (8, 9, 10, 14, 16):
            invoices.cell(row_index, col_index).number_format = "#,##0.00"
        match_cell = invoices.cell(row_index, 12)
        match_cell.font = Font(
            name="Arial", bold=True,
            color=green if item.get("exact_reference_match") else red,
        )
        invoices.row_dimensions[row_index].height = 23

    widths = [18, 16, 18, 28, 16, 18, 12, 18, 15, 15, 18,
              16, 18, 18, 25, 16]
    for index, width in enumerate(widths, 1):
        invoices.column_dimensions[get_column_letter(index)].width = width

    summary = workbook.create_sheet("الملخص", 0)
    summary.sheet_view.rightToLeft = True
    summary.merge_cells("A1:D1")
    summary["A1"] = "ملخص مقارنة سلة مع فواتير قيود"
    summary["A1"].fill = PatternFill("solid", fgColor=navy)
    summary["A1"].font = Font(name="Arial", color=white, bold=True, size=16)
    summary["A1"].alignment = Alignment(horizontal="right")
    summary.row_dimensions[1].height = 32
    summary_rows = [
        ("من تاريخ", report.get("from_date")),
        ("إلى تاريخ", report.get("to_date")),
        ("البحث", report.get("search") or "الكل"),
        (
            "طلبات سلة المؤهلة",
            report["summary"]["eligible_salla_orders"],
        ),
        ("فواتير قيود", report["summary"]["qoyod_invoices"]),
        (
            "مراجع مطابقة تمامًا",
            report["summary"]["exact_reference_matches"],
        ),
        ("فواتير بلا مرجع سلة مطابق",
         report["summary"]["unmatched_qoyod_invoices"]),
        ("صفوف التقرير بعد البحث", report.get("total", 0)),
        ("آخر مزامنة", report["summary"].get("latest_sync_at") or "—"),
    ]
    for row_index, (label, value) in enumerate(summary_rows, 3):
        summary.cell(row_index, 1, label)
        summary.cell(row_index, 2, _excel_safe_value(value))
        for col_index in (1, 2):
            cell = summary.cell(row_index, col_index)
            cell.border = border
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(name="Arial", bold=(col_index == 1), size=11)
            if col_index == 1:
                cell.fill = PatternFill("solid", fgColor=pale)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 4
    summary.column_dimensions["D"].width = 4

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def safe_export_filename(from_date: str, to_date: str) -> str:
    """ASCII-only filename, safe for Content-Disposition."""
    clean_from = re.sub(r"[^0-9-]", "", from_date)
    clean_to = re.sub(r"[^0-9-]", "", to_date)
    return f"mezan_qoyod_invoices_{clean_from}_{clean_to}.xlsx"
