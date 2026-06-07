"""Iter-85 — Orders Explorer endpoints."""
import os
import requests
import pytest

BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    open("/app/frontend/.env").read().split("REACT_APP_BACKEND_URL=")[1].split("\n")[0],
)
EMAIL = "amasi.jewelery@gmail.com"
PASSWORD = "10201917"


@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": EMAIL, "password": PASSWORD}, timeout=10)
    r.raise_for_status()
    s.headers.update({"Authorization": f"Bearer {r.json()['access_token']}"})
    return s


def test_status_summary_returns_rows_and_categories(auth):
    r = auth.get(f"{BASE_URL}/api/orders/status-summary", timeout=15)
    r.raise_for_status()
    d = r.json()
    assert "rows" in d and "by_category" in d and "totals" in d
    # Real merchant has at least 10 statuses
    assert len(d["rows"]) >= 10
    # 4 categories present
    for c in ("confirmed", "pending", "refunded", "cancelled"):
        assert c in d["by_category"]
        assert "count" in d["by_category"][c]
        assert "amount" in d["by_category"][c]
    # Totals match sum of category counts
    by_cat_sum = sum(v["count"] for v in d["by_category"].values())
    assert d["totals"]["orders_count"] == by_cat_sum


def test_status_summary_well_known_buckets(auth):
    d = auth.get(f"{BASE_URL}/api/orders/status-summary", timeout=15).json()
    by_status = {r["status"]: r for r in d["rows"]}
    # User-mentioned statuses must be present
    for s in ("تم التوصيل", "تم التنفيذ", "جاري التوصيل", "قيد التنفيذ"):
        assert s in by_status, by_status
    # تم التوصيل should be confirmed by default
    assert by_status["تم التوصيل"]["category"] == "confirmed"
    assert by_status["جاري التوصيل"]["category"] == "pending"


def test_orders_list_paginated(auth):
    r = auth.get(f"{BASE_URL}/api/orders?limit=10&page=1", timeout=15).json()
    assert r["page"] == 1
    assert r["limit"] == 10
    assert isinstance(r["items"], list)
    assert len(r["items"]) <= 10
    assert r["total"] >= 2000  # real merchant
    # each item has category
    for o in r["items"]:
        assert "category" in o


def test_orders_filter_by_status(auth):
    r = auth.get(
        f"{BASE_URL}/api/orders?status=%D8%AA%D9%85%20%D8%A7%D9%84%D8%AA%D9%88%D8%B5%D9%8A%D9%84&limit=5",
        timeout=15,
    ).json()
    assert r["total"] >= 1000  # تم التوصيل dominates the dataset
    for o in r["items"]:
        assert o.get("order_status") == "تم التوصيل"
        assert o["category"] == "confirmed"


def test_orders_filter_by_date_range(auth):
    r = auth.get(
        f"{BASE_URL}/api/orders?from_date=2026-05-01&to_date=2026-05-31&limit=5",
        timeout=15,
    ).json()
    for o in r["items"]:
        d = o.get("order_date") or ""
        assert "2026-05" in d, o


def test_orders_search(auth):
    # grab first order to search by its number
    first = auth.get(f"{BASE_URL}/api/orders?limit=1", timeout=15).json()["items"][0]
    num = first["order_number"]
    r = auth.get(f"{BASE_URL}/api/orders?search={num}&limit=5", timeout=15).json()
    assert r["total"] >= 1
    assert any(o.get("order_number") == num for o in r["items"])


def test_summary_respects_date_window(auth):
    full = auth.get(f"{BASE_URL}/api/orders/status-summary", timeout=15).json()
    narrow = auth.get(
        f"{BASE_URL}/api/orders/status-summary?from_date=2026-05-01&to_date=2026-05-31",
        timeout=15,
    ).json()
    # narrow window must be ≤ full
    assert narrow["totals"]["orders_count"] <= full["totals"]["orders_count"]


def test_category_filter_matches_summary_buckets(auth):
    """Iter-85 regression: clicking a category card must return EXACTLY
    the same count as the corresponding summary bucket."""
    s = auth.get(f"{BASE_URL}/api/orders/status-summary", timeout=15).json()
    for cat in ("confirmed", "pending", "refunded", "cancelled"):
        expected = s["by_category"][cat]["count"]
        r = auth.get(
            f"{BASE_URL}/api/orders?category={cat}&limit=5", timeout=15,
        ).json()
        assert r["total"] == expected, (cat, r["total"], expected)
        # every returned item must belong to that category
        for o in r["items"]:
            assert o["category"] == cat, (cat, o)


def test_combined_filters_dont_clobber(auth):
    """status + search together should AND-combine (not lose one)."""
    r = auth.get(
        f"{BASE_URL}/api/orders?status=%D8%AA%D9%85%20%D8%A7%D9%84%D8%AA%D9%88%D8%B5%D9%8A%D9%84&search=2643&limit=5",
        timeout=15,
    ).json()
    for o in r["items"]:
        assert o["order_status"] == "تم التوصيل"
        assert "2643" in (o.get("order_number") or "") \
            or "2643" in (o.get("customer_name") or "") \
            or "2643" in (o.get("customer_phone") or ""), o


def test_export_xlsx_returns_filtered_workbook(auth, tmp_path):
    """Iter-86 — export honours filters and contains both sheets."""
    from openpyxl import load_workbook
    r = auth.get(
        f"{BASE_URL}/api/orders/export.xlsx?category=cancelled",
        timeout=20,
    )
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    p = tmp_path / "out.xlsx"
    p.write_bytes(r.content)
    wb = load_workbook(p)
    assert "الطلبات" in wb.sheetnames
    assert "ملخّص" in wb.sheetnames
    ws = wb["الطلبات"]
    # 59 cancelled orders + 1 header row
    assert ws.max_row == 60
    # all rows are cancelled
    for row_idx in range(2, ws.max_row + 1):
        assert ws.cell(row_idx, 9).value == "ملغاة"
