"""Iter-74 (Phase 80) — Payment-gateway settlement imports.

Tests verify the 3 parsers against REAL sample files we downloaded
from the merchant's Google Drive, the dedup contract, the order
matching + actual_* field population, the rollback on delete, and
the coverage analytics endpoint.
"""
import os
import sys

import pytest
import openpyxl
import requests

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from settlements_import.registry import detect_provider, parse  # noqa: E402
from settlements_import.service import _normalize_order_number, _consolidate_rows  # noqa: E402


BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    "https://salla-analytics.preview.emergentagent.com",
).rstrip("/")
EMAIL = "amasi.jewelery@gmail.com"
PASSWORD = "10201917"
SAMPLES = "/tmp/settlements_samples"


@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=15)
    assert r.status_code == 200, r.text
    s.headers.update({"Authorization": f"Bearer {r.json()['access_token']}"})
    return s


@pytest.fixture(autouse=True)
def _clean(auth):
    """Wipe any prior uploads from this user before each test to keep
    counts deterministic."""
    r = auth.get(f"{BASE_URL}/api/payment-settlements", timeout=10)
    for f in r.json().get("files", []):
        auth.delete(f"{BASE_URL}/api/payment-settlements/{f['id']}", timeout=10)
    yield
    r = auth.get(f"{BASE_URL}/api/payment-settlements", timeout=10)
    for f in r.json().get("files", []):
        auth.delete(f"{BASE_URL}/api/payment-settlements/{f['id']}", timeout=10)


# ── 1. Provider detection ─────────────────────────────────────────────
def test_detect_provider_salla():
    wb = openpyxl.load_workbook(f"{SAMPLES}/salla.xlsx", data_only=True)
    assert detect_provider(wb) == "salla"
    wb.close()


def test_detect_provider_tamara():
    wb = openpyxl.load_workbook(f"{SAMPLES}/tamara.xlsx", data_only=True)
    assert detect_provider(wb) == "tamara"
    wb.close()


def test_detect_provider_tabby():
    wb = openpyxl.load_workbook(f"{SAMPLES}/tabby.xlsx", data_only=True)
    assert detect_provider(wb) == "tabby"
    wb.close()


# ── 2. Salla parser produces the expected totals ──────────────────────
def test_salla_parser_totals():
    wb = openpyxl.load_workbook(f"{SAMPLES}/salla.xlsx", data_only=True)
    try:
        res = parse("salla", wb)
    finally:
        wb.close()
    assert res["provider"] == "salla"
    assert res["totals"]["rows"] == 140
    assert abs(res["totals"]["gross"] - 26686.32) < 0.05
    assert abs(res["totals"]["fees"] - 500.38) < 0.05
    assert abs(res["totals"]["fees_vat"] - 74.97) < 0.05
    assert abs(res["totals"]["net"] - 26110.97) < 0.05
    # Header captures invoice number
    assert res["header"]["invoice_number"] == "6320306"
    # Sample row maps mada → mada and BTC card → credit_card
    sample = res["entries"][0]
    assert sample["actual_payment_method"] == "mada"
    assert sample["event_type"] == "sale"
    assert sample["actual_fee_rate"] > 0


# ── 3. Tamara parser handles refunds + statement metadata ─────────────
def test_tamara_parser_totals_and_refunds():
    wb = openpyxl.load_workbook(f"{SAMPLES}/tamara.xlsx", data_only=True)
    try:
        res = parse("tamara", wb)
    finally:
        wb.close()
    assert res["provider"] == "tamara"
    assert res["totals"]["rows"] == 131
    assert abs(res["totals"]["gross"] - 25213.54) < 0.05
    assert abs(res["totals"]["refund_full"] - 1581.20) < 0.05
    assert abs(res["totals"]["refund_partial"] - 128.36) < 0.05
    # Statement metadata
    assert res["header"]["statement_id"] == "P0420741SA260606"
    assert "Tamara Merchant ID" in str(res["header"].get("tamara_merchant_id", "")) or len(res["header"].get("tamara_merchant_id", "")) > 30
    # At least one refund entry exists
    refunds = [e for e in res["entries"] if e["event_type"] == "refund"]
    assert len(refunds) >= 2
    # Refund row never carries positive fees
    for r in refunds:
        assert r["actual_payment_fee"] == 0.0
        assert r["actual_payment_vat"] == 0.0
        assert r["actual_net_amount"] <= 0


# ── 4. Tabby parser totals + sale/refund split ────────────────────────
def test_tabby_parser_totals():
    wb = openpyxl.load_workbook(f"{SAMPLES}/tabby.xlsx", data_only=True)
    try:
        res = parse("tabby", wb)
    finally:
        wb.close()
    assert res["provider"] == "tabby"
    assert res["totals"]["rows"] == 82
    assert abs(res["totals"]["gross"] - 15771.96) < 0.05
    assert abs(res["totals"]["net"] - 13815.78) < 0.05
    assert res["header"]["statement_id"] == "Tabby20260601SAR"
    sample = res["entries"][0]
    assert sample["actual_payment_method"] == "tabby"
    assert sample["actual_fee_rate"] == 6.99


# ── 5. Order-number normalization edge cases ──────────────────────────
def test_normalize_order_number_strips_decimals():
    assert _normalize_order_number("263724404") == "263724404"
    assert _normalize_order_number(263724404) == "263724404"
    assert _normalize_order_number("263724404.0") == "263724404"
    assert _normalize_order_number(263724404.0) == "263724404"
    assert _normalize_order_number(None) == ""
    assert _normalize_order_number("") == ""
    assert _normalize_order_number("  ABC-123  ") == "ABC-123"


# ── 6. Consolidate sale + refund rows for same order ──────────────────
def test_consolidate_capture_plus_refund():
    rows = [
        {"actual_gross_amount": 245.77, "actual_payment_fee": 16.18, "actual_payment_vat": 2.43,
         "actual_net_amount": 227.16, "actual_refund_amount": 0, "actual_partial_refund_amount": 0,
         "actual_fee_rate": 6.99, "actual_payment_method": "tamara",
         "settlement_date": "2026-05-15", "settlement_reference": "STMT-1"},
        {"actual_gross_amount": 245.77, "actual_payment_fee": 0, "actual_payment_vat": 0,
         "actual_net_amount": -128.36, "actual_refund_amount": 0, "actual_partial_refund_amount": 128.36,
         "actual_fee_rate": 0, "actual_payment_method": "tamara",
         "settlement_date": "2026-06-02", "settlement_reference": "STMT-1"},
    ]
    out = _consolidate_rows(rows)
    af = out["actual_fields"]
    assert af["actual_gross_amount"] == 245.77
    assert af["actual_payment_fee"] == 16.18
    assert af["actual_net_amount"] == round(227.16 - 128.36, 2)
    assert af["actual_partial_refund_amount"] == 128.36
    assert af["actual_refund_amount"] == 0
    # latest event date wins
    assert out["settlement_date"] == "2026-06-02"


# ── 7. Upload + match (end-to-end on REAL files) ──────────────────────
def test_upload_salla_matches_real_orders(auth):
    with open(f"{SAMPLES}/salla.xlsx", "rb") as fh:
        r = auth.post(
            f"{BASE_URL}/api/payment-settlements/upload",
            files={"file": ("salla.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "imported"
    assert body["provider"] == "salla"
    # The merchant's real data should match at least 100 of 140 rows
    assert body["matched"] >= 100
    assert body["totals"]["net"] > 25000


def test_upload_tamara_picks_up_metadata(auth):
    with open(f"{SAMPLES}/tamara.xlsx", "rb") as fh:
        r = auth.post(
            f"{BASE_URL}/api/payment-settlements/upload",
            files={"file": ("tamara.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["provider"] == "tamara"
    assert body["header"]["statement_id"] == "P0420741SA260606"
    assert body["totals"]["refund_full"] > 0


def test_upload_tabby_matches(auth):
    with open(f"{SAMPLES}/tabby.xlsx", "rb") as fh:
        r = auth.post(
            f"{BASE_URL}/api/payment-settlements/upload",
            files={"file": ("tabby.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["provider"] == "tabby"
    assert body["matched"] >= 50


# ── 8. Dedup: re-uploading same file returns 'duplicate' ──────────────
def test_dedup_same_file_twice(auth):
    with open(f"{SAMPLES}/salla.xlsx", "rb") as fh:
        first = auth.post(
            f"{BASE_URL}/api/payment-settlements/upload",
            files={"file": ("salla.xlsx", fh, "application/vnd.ms-excel")},
            timeout=30,
        )
    assert first.status_code == 200 and first.json()["status"] == "imported"
    file_id = first.json()["file_id"]

    with open(f"{SAMPLES}/salla.xlsx", "rb") as fh:
        second = auth.post(
            f"{BASE_URL}/api/payment-settlements/upload",
            files={"file": ("salla.xlsx", fh, "application/vnd.ms-excel")},
            timeout=30,
        )
    assert second.status_code == 200
    assert second.json()["status"] == "duplicate"
    assert second.json()["file_id"] == file_id


# ── 9. Coverage analytics reflects uploaded files ─────────────────────
def test_coverage_analytics_after_upload(auth):
    with open(f"{SAMPLES}/salla.xlsx", "rb") as fh:
        auth.post(
            f"{BASE_URL}/api/payment-settlements/upload",
            files={"file": ("salla.xlsx", fh, "application/vnd.ms-excel")},
            timeout=30,
        )
    r = auth.get(f"{BASE_URL}/api/payment-settlements/_analytics/coverage", timeout=15)
    assert r.status_code == 200
    body = r.json()
    assert "totals" in body and "actual_aggregates" in body and "by_provider" in body
    assert body["totals"]["orders_actual"] > 0
    salla = next((p for p in body["by_provider"] if p["provider"] == "salla"), None)
    assert salla is not None
    assert salla["orders"] > 0


# ── 10. Delete file rolls back actual_* fields ────────────────────────
def test_delete_rolls_back_actual_fields(auth):
    with open(f"{SAMPLES}/tabby.xlsx", "rb") as fh:
        r = auth.post(
            f"{BASE_URL}/api/payment-settlements/upload",
            files={"file": ("tabby.xlsx", fh, "application/vnd.ms-excel")},
            timeout=30,
        )
    file_id = r.json()["file_id"]
    matched = r.json()["matched"]
    assert matched > 0

    r = auth.delete(f"{BASE_URL}/api/payment-settlements/{file_id}", timeout=15)
    assert r.status_code == 200
    body = r.json()
    assert body["removed"] == 1
    assert body["orders_rolled_back"] == matched

    # Coverage should drop back to 0 actual orders attributable to tabby
    r = auth.get(f"{BASE_URL}/api/payment-settlements/_analytics/coverage", timeout=10)
    tabby = next((p for p in r.json()["by_provider"] if p["provider"] == "tabby"), None)
    assert tabby is None or tabby["orders"] == 0


# ── 11. Bad file rejected with Arabic error ───────────────────────────
def test_non_xlsx_rejected(auth):
    r = auth.post(
        f"{BASE_URL}/api/payment-settlements/upload",
        files={"file": ("test.txt", b"not an xlsx file", "text/plain")},
        timeout=10,
    )
    assert r.status_code == 400
    assert "تعذّر فتح" in r.json()["detail"] or "تعذر فتح" in r.json()["detail"]


def test_empty_file_rejected(auth):
    r = auth.post(
        f"{BASE_URL}/api/payment-settlements/upload",
        files={"file": ("empty.xlsx", b"", "application/vnd.ms-excel")},
        timeout=10,
    )
    assert r.status_code == 400
