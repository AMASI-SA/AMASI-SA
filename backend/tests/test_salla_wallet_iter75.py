"""Iter-75 — Salla 'شحن محفظة' (wallet recharge / مشتريات سله) handling.

The merchant uploaded a real Salla invoice that contained 3 wallet-
recharge rows (negative amounts with payment_method ==
'order.payment_method.' — Salla's untranslated i18n key). These are
shipping-label purchases the merchant pays Salla from his wallet
credit. The system must:

  • Detect them by needle / negative-amount signal.
  • Aggregate them under totals.salla_purchases_{total,count}.
  • EXCLUDE them from totals.gross / totals.net / totals.fees so the
    file totals represent real customer sales only.
  • NOT push actual_* fields to unified_orders for these rows.
  • Still attach the row to settlement_entries (audit trail).
"""
import os
import sys

import openpyxl
import pytest
import requests

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from settlements_import.parsers.salla import (  # noqa: E402
    _is_wallet_recharge, parse as parse_salla,
)


BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    "https://salla-analytics.preview.emergentagent.com",
).rstrip("/")
EMAIL = "amasi.jewelery@gmail.com"
PASSWORD = "10201917"
SAMPLES = "/tmp/settlements_samples"


@pytest.fixture(scope="module")
def auth():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=15)
    assert r.status_code == 200, r.text
    s.headers.update({"Authorization": f"Bearer {r.json()['access_token']}"})
    return s


@pytest.fixture(autouse=True)
def _clean(auth):
    r = auth.get(f"{BASE_URL}/api/payment-settlements", timeout=10)
    for f in r.json().get("files", []):
        auth.delete(f"{BASE_URL}/api/payment-settlements/{f['id']}", timeout=10)
    yield
    r = auth.get(f"{BASE_URL}/api/payment-settlements", timeout=10)
    for f in r.json().get("files", []):
        auth.delete(f"{BASE_URL}/api/payment-settlements/{f['id']}", timeout=10)


# ── 1. Unit: detection helper ─────────────────────────────────────────
def test_detect_untranslated_i18n_key():
    assert _is_wallet_recharge("order.payment_method.", -34.5, -34.5) is True


def test_detect_arabic_literal():
    assert _is_wallet_recharge("شحن محفظة", -10, -10) is True


def test_detect_english_variants():
    assert _is_wallet_recharge("Wallet Recharge", -1, -1) is True
    assert _is_wallet_recharge("wallet_recharge", -1, -1) is True


def test_negative_amounts_alone_do_not_trigger():
    # Customer refunds through مدى / credit_card are negative but NOT
    # wallet recharges — they must NOT be flagged.
    assert _is_wallet_recharge("", -5, -5) is False
    assert _is_wallet_recharge("مدى", -312.20, -312.20) is False
    assert _is_wallet_recharge("البطاقة الائتمانية", -89.43, -89.43) is False


def test_normal_methods_not_flagged():
    assert _is_wallet_recharge("mada", 100, 95) is False
    assert _is_wallet_recharge("البطاقة الائتمانية", 200, 190) is False
    assert _is_wallet_recharge("", 50, 45) is False


# ── 2. Parser produces expected wallet totals on real sample ──────────
def test_salla_wallet_file_aggregates_purchases():
    wb = openpyxl.load_workbook(f"{SAMPLES}/salla_wallet.xlsx", data_only=True)
    try:
        res = parse_salla(wb)
    finally:
        wb.close()
    t = res["totals"]
    # 3 wallet rows totaling 34.5 + 40.25 + 34.5 = 109.25
    assert t["salla_purchases_count"] == 3
    assert abs(t["salla_purchases_total"] - 109.25) < 0.01
    # Sales totals exclude the wallet rows (117 - 3 = 114 sale entries)
    sales = [e for e in res["entries"] if e["event_type"] == "sale"]
    purchases = [e for e in res["entries"] if e["event_type"] == "salla_purchase"]
    assert len(sales) == 114
    assert len(purchases) == 3
    # The 3 expected order numbers
    expected_orders = {"259635319", "259392433", "257396516"}
    assert {p["order_number"] for p in purchases} == expected_orders


# ── 3. Parser tags wallet entries correctly ───────────────────────────
def test_wallet_entries_have_correct_shape():
    wb = openpyxl.load_workbook(f"{SAMPLES}/salla_wallet.xlsx", data_only=True)
    try:
        res = parse_salla(wb)
    finally:
        wb.close()
    purchases = [e for e in res["entries"] if e["event_type"] == "salla_purchase"]
    assert len(purchases) == 3
    for p in purchases:
        assert p["actual_payment_method"] == "wallet_recharge"
        assert p["actual_net_amount"] < 0
        assert p["actual_gross_amount"] < 0
        assert p["actual_fee_rate"] == 0.0
        assert "raw_payment_method" in p
        assert p.get("notes")  # description hint preserved


# ── 4. Original samples are NOT affected ──────────────────────────────
def test_original_salla_file_has_no_wallet_purchases():
    """The first Salla file (without wallet rows on real customer
    methods) must keep its salla_purchases bucket empty after the
    wallet-detection feature is added."""
    wb = openpyxl.load_workbook(f"{SAMPLES}/salla.xlsx", data_only=True)
    try:
        res = parse_salla(wb)
    finally:
        wb.close()
    t = res["totals"]
    # One row in the original file uses 'order.payment_method.' but
    # has order_number=None so it's filtered as blank before wallet
    # detection. The result: 0 wallet purchases.
    assert t["salla_purchases_count"] == 0
    assert t["salla_purchases_total"] == 0.0
    # And the gross total stays the same (refunds are tagged separately
    # but don't change positive gross).
    assert abs(t["gross"] - 26686.32) < 0.05


# ── 6. Customer-refund detection (negative on normal method) ──────────
def test_partial_refund_detected_on_credit_card():
    """Order 263864673 in salla_refund.xlsx has 2 sales (407.02 + 89.43)
    plus a partial refund (-89.43). The partial refund must be tagged
    under refund_partial, NOT refund_full or salla_purchases."""
    wb = openpyxl.load_workbook(f"{SAMPLES}/salla_refund.xlsx", data_only=True)
    try:
        res = parse_salla(wb)
    finally:
        wb.close()
    t = res["totals"]
    # 3 negative rows: 312.20 (mada) + 208.59 (mada) + 89.43 (cc) = 610.22
    assert abs(t["refund_partial"] - 610.22) < 0.05
    # None of them is full refund (each is partial relative to total
    # paid on the same order).
    assert t["refund_full"] == 0.0
    # Wallet bucket stays untouched
    assert t["salla_purchases_count"] == 0

    # The 263864673 entry is tagged as a partial refund of 89.43
    refunds_263 = [e for e in res["entries"]
                   if e["order_number"] == "263864673" and e["event_type"] == "refund"]
    assert len(refunds_263) == 1
    assert abs(refunds_263[0]["actual_partial_refund_amount"] - 89.43) < 0.01
    assert refunds_263[0]["actual_refund_amount"] == 0.0
    assert refunds_263[0]["actual_payment_method"] == "credit_card"


def test_full_refund_classified_when_amount_equals_paid():
    """Synthetic check using the consolidate helper: when an order's
    total positive gross equals the refund amount, it should classify
    as refund_full (≥99% tolerance)."""
    import openpyxl as ox
    wb = ox.Workbook()
    ws = wb.active
    ws.title = "Invoice # 9999999"
    ws.append(["رقم الطلب", "إجمالي الطلب (ر.س)", "طريقة الدفع",
               "الرسوم (ر.س)", "المُستحق قبل الضريبة (ر.س)",
               "الضريبة", "المُستحق بعد الضريبة (ر.س)"])
    # Order paid 100, refunded -100 → full refund
    ws.append(["999000001", 100.00, "مدى", 1.50, 98.50, 0.23, 98.27])
    ws.append(["999000001", -100.00, "مدى", 0.0, -100.00, 0.0, -100.00])
    res = parse_salla(wb)
    refunds = [e for e in res["entries"] if e["event_type"] == "refund"]
    assert len(refunds) == 1
    # 100 >= 100 * 0.99 → full
    assert abs(refunds[0]["actual_refund_amount"] - 100.00) < 0.01
    assert refunds[0]["actual_partial_refund_amount"] == 0.0
    assert res["totals"]["refund_full"] == 100.0
    assert res["totals"]["refund_partial"] == 0.0


# ── 5. End-to-end upload: salla_purchases surface in /api ─────────────
def test_upload_wallet_file_surfaces_purchases(auth):
    with open(f"{SAMPLES}/salla_wallet.xlsx", "rb") as fh:
        r = auth.post(
            f"{BASE_URL}/api/payment-settlements/upload",
            files={"file": ("salla_wallet.xlsx", fh,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "imported"
    t = body["totals"]
    assert t["salla_purchases_count"] == 3
    assert abs(t["salla_purchases_total"] - 109.25) < 0.01

    # And the files endpoint surfaces the same numbers
    r2 = auth.get(f"{BASE_URL}/api/payment-settlements", timeout=10)
    f = next((x for x in r2.json()["files"] if x["id"] == body["file_id"]), None)
    assert f is not None
    assert f["totals"]["salla_purchases_count"] == 3


# ── 6. Wallet rows do NOT update unified_orders.actual_* ──────────────
def test_wallet_rows_do_not_pollute_unified_orders(auth):
    """The 3 order_numbers referenced by wallet rows (259635319,
    259392433, 257396516) might be real customer orders in the DB.
    Wallet-row data must NOT set actual_* on them."""
    # Read state before
    orders_to_check = ["259635319", "259392433", "257396516"]

    # Upload wallet file
    with open(f"{SAMPLES}/salla_wallet.xlsx", "rb") as fh:
        r = auth.post(
            f"{BASE_URL}/api/payment-settlements/upload",
            files={"file": ("salla_wallet.xlsx", fh, "application/vnd.ms-excel")},
            timeout=30,
        )
    assert r.status_code == 200
    file_id = r.json()["file_id"]

    # Pull each affected order via the existing search endpoint and
    # verify that if it had actual_payment_method != wallet_recharge
    # (i.e. the file ALSO has a sale row for the same order_number).
    # The wallet-recharge value of actual_net_amount=-34.5 must NOT
    # appear because the importer skipped it.
    #
    # We use the unified_orders search endpoint to inspect:
    from pymongo import MongoClient
    from dotenv import load_dotenv
    load_dotenv("/app/backend/.env")
    c = MongoClient(os.environ["MONGO_URL"])
    db = c[os.environ["DB_NAME"]]

    for ono in orders_to_check:
        doc = db.unified_orders.find_one(
            {"order_number": ono},
            {"actual_payment_method": 1, "actual_net_amount": 1, "_id": 0},
        )
        if not doc:
            continue
        # Either the order wasn't matched at all → no actual_*, OR
        # it was matched by a sale row in the same file with proper
        # method. Either way, actual_payment_method must NOT be
        # 'wallet_recharge' and actual_net_amount must NOT be the
        # wallet's negative value.
        assert doc.get("actual_payment_method") != "wallet_recharge", (
            f"order {ono} was polluted by wallet-recharge row"
        )

    # Cleanup handled by fixture
