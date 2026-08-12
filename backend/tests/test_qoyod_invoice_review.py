"""Focused tests for the standalone read-only Qoyod invoice review."""
from __future__ import annotations

import asyncio
from datetime import date
from io import BytesIO

from fastapi import APIRouter, FastAPI
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
import pytest

from integrations.qoyod.invoice_review import (
    _excel_safe_value,
    build_invoice_review,
    build_invoice_review_workbook,
    parse_review_range,
    sync_invoice_review,
)
from integrations.qoyod.invoice_review_routes import (
    attach_invoice_review_routes,
)


def _value(row, key):
    current = row
    for part in key.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


def _matches(row, query):
    for key, expected in query.items():
        actual = _value(row, key)
        if isinstance(expected, dict):
            if "$in" in expected and actual not in expected["$in"]:
                return False
            if "$gte" in expected and (
                actual is None or actual < expected["$gte"]
            ):
                return False
            if "$lte" in expected and (
                actual is None or actual > expected["$lte"]
            ):
                return False
            continue
        if actual != expected:
            return False
    return True


class _Cursor:
    def __init__(self, rows):
        self.rows = list(rows)

    def sort(self, spec, direction=None):
        if isinstance(spec, list):
            fields = spec
        else:
            fields = [(spec, direction or 1)]
        for field, order in reversed(fields):
            self.rows.sort(
                key=lambda row: str(_value(row, field) or ""),
                reverse=order < 0,
            )
        return self

    def __aiter__(self):
        self._iterator = iter(self.rows)
        return self

    async def __anext__(self):
        try:
            return next(self._iterator)
        except StopIteration as exc:
            raise StopAsyncIteration from exc


class _Collection:
    def __init__(self, rows):
        self.rows = list(rows)
        self.find_queries = []

    def find(self, query, _projection=None):
        self.find_queries.append(query)
        return _Cursor([
            dict(row) for row in self.rows if _matches(row, query)
        ])


class _DB:
    def __init__(self, *, orders=(), invoices=()):
        self.unified_orders = _Collection(orders)
        self.qoyod_invoices = _Collection(invoices)


def _run(awaitable):
    return asyncio.run(awaitable)


def _order(
    owner, number, status, *, order_date="2026-08-10", total=100,
    order_date_inferred=False,
):
    return {
        "user_id": owner,
        "order_number": number,
        "order_date": order_date,
        "order_date_inferred": order_date_inferred,
        "order_status_slug": status,
        "order_status": status,
        "total_amount": total,
        "customer_name": f"Salla {number}",
    }


def _invoice(
    owner,
    invoice_id,
    reference,
    *,
    issue_date="2026-08-12",
    customer="عميل قيود",
    invoice_number=None,
    total=100,
    due_date="2026-08-20",
    currency="SAR",
):
    return {
        "user_id": owner,
        "qoyod_invoice_id": invoice_id,
        "invoice_number": invoice_number or f"INV-{invoice_id}",
        "reference": reference,
        "customer_name": customer,
        "issue_date": issue_date,
        "total": total,
        "paid_amount": total,
        "remaining": 0,
        "status": "paid",
        "last_sync_at": "2026-08-12T11:00:00+00:00",
        "raw_response": {
            "due_date": due_date,
            "currency": currency,
        },
    }


def _review_db():
    return _DB(
        orders=[
            _order("merchant-1", "277000001", "completed", total=100),
            _order("merchant-1", "277000002", "in_delivery", total=200),
            _order("merchant-1", "277000003", "delivered", total=300),
            _order("merchant-1", "277000004", "processing", total=400),
            _order("merchant-1", "277000005", "مكتمل", total=500),
            _order("merchant-1", "277000006", "completed", total=600,
                   order_date_inferred=True),
            _order("merchant-1", "276000000", "completed",
                   order_date="2026-06-30"),
            _order("merchant-2", "277000099", "completed"),
        ],
        invoices=[
            _invoice("main", "1401", "277000001", total=100),
            _invoice("main", "1402", "NO-SALLA-REFERENCE",
                     customer="عميلة مستقلة", invoice_number="Q-900"),
            _invoice("main", "1405", "277000004", total=400),
            _invoice("main", "1406", "276000000", total=100),
            _invoice("main", "DRY:1403", "277000002"),
            _invoice("main", "1404", "277000003",
                     issue_date="2026-06-30"),
            _invoice("another-qoyod-tenant", "9999", "277000002"),
        ],
    )


def test_review_floor_and_invalid_range():
    assert parse_review_range(
        "2026-01-01", "2026-08-12", today=date(2026, 8, 12)
    ) == (date(2026, 7, 1), date(2026, 8, 12))
    with pytest.raises(ValueError, match="from_date_must_not_be_after"):
        parse_review_range("2026-08-12", "2026-08-11")


def test_list_counts_three_eligible_statuses_and_exact_reference_only():
    report = _run(build_invoice_review(
        _review_db(),
        orders_user_id="merchant-1",
        markers_user_id="main",
        from_date="2026-07-01",
        to_date="2026-08-12",
    ))

    assert report["page_size"] == 15
    assert report["summary"] == {
        "eligible_salla_orders": 3,
        "qoyod_invoices": 4,
        "qoyod_distinct_references": 4,
        "eligible_with_qoyod_invoice": 2,
        "eligible_without_qoyod_invoice": 1,
        "qoyod_outside_eligible": 3,
        "exact_reference_matches": 3,
        "unmatched_qoyod_invoices": 1,
        "latest_sync_at": "2026-08-12T11:00:00+00:00",
    }
    assert report["last_sync_at"] == "2026-08-12T11:00:00+00:00"
    assert report["total"] == 4
    matched = next(
        item for item in report["items"]
        if item["reference"] == "277000001"
    )
    assert matched["exact_reference_match"] is True
    assert matched["salla_order_number"] == "277000001"
    assert matched["salla_status"] == "completed"
    assert matched["salla_total"] == 100.0
    assert matched["due_date"] == "2026-08-20"
    assert matched["currency"] == "SAR"
    unmatched = next(
        item for item in report["items"]
        if item["reference"] == "NO-SALLA-REFERENCE"
    )
    assert unmatched["exact_reference_match"] is False
    assert unmatched["salla_status"] is None
    processing = next(
        item for item in report["items"]
        if item["reference"] == "277000004"
    )
    assert processing["exact_reference_match"] is True
    assert processing["salla_status"] == "processing"
    historical = next(
        item for item in report["items"]
        if item["reference"] == "276000000"
    )
    assert historical["exact_reference_match"] is True
    # Neither processing, a non-approved alias, an inferred date, nor the
    # pre-floor order inflates eligibility.  The real invoice dated before
    # this report window still proves that eligible order 277000003 was sent.
    assert report["summary"]["eligible_salla_orders"] == 3


def test_comparable_counts_use_distinct_references_not_invoice_rows():
    db = _DB(
        orders=[_order("merchant-1", "277000001", "completed")],
        invoices=[
            _invoice("main", "1401", "277000001"),
            _invoice("main", "1402", "277000001"),
            _invoice("main", "1403", None),
        ],
    )
    report = _run(build_invoice_review(
        db,
        orders_user_id="merchant-1",
        markers_user_id="main",
        from_date="2026-07-01",
        to_date="2026-08-12",
    ))

    assert report["summary"] == {
        "eligible_salla_orders": 1,
        "qoyod_invoices": 3,
        "qoyod_distinct_references": 1,
        "eligible_with_qoyod_invoice": 1,
        "eligible_without_qoyod_invoice": 0,
        "qoyod_outside_eligible": 0,
        "exact_reference_matches": 2,
        "unmatched_qoyod_invoices": 1,
        "latest_sync_at": "2026-08-12T11:00:00+00:00",
    }


def test_salla_queries_are_scoped_to_range_and_invoice_references():
    db = _review_db()
    _run(build_invoice_review(
        db,
        orders_user_id="merchant-1",
        markers_user_id="main",
        from_date="2026-07-01",
        to_date="2026-08-12",
    ))

    assert db.unified_orders.find_queries == [
        {
            "user_id": "merchant-1",
            "order_date": {
                "$gte": "2026-07-01",
                "$lte": "2026-08-12",
            },
        },
        {
            "user_id": "merchant-1",
            "order_number": {
                "$in": [
                    "276000000", "277000004", "NO-SALLA-REFERENCE",
                ],
            },
        },
    ]


@pytest.mark.parametrize(
    ("search", "expected_reference"),
    [
        ("277000001", "277000001"),
        ("Q-900", "NO-SALLA-REFERENCE"),
        ("مستقلة", "NO-SALLA-REFERENCE"),
    ],
)
def test_searches_reference_invoice_number_and_customer(
    search, expected_reference,
):
    report = _run(build_invoice_review(
        _review_db(),
        orders_user_id="merchant-1",
        markers_user_id="main",
        from_date="2026-07-01",
        to_date="2026-08-12",
        search=search,
    ))
    assert report["total"] == 1
    assert report["items"][0]["reference"] == expected_reference
    # Summary is the selected date range, independent of search.
    assert report["summary"]["qoyod_invoices"] == 4


def test_pagination_is_fifteen_capable():
    db = _DB(
        orders=[],
        invoices=[
            _invoice("main", str(1500 + i), f"2771{i:05d}")
            for i in range(17)
        ],
    )
    first = _run(build_invoice_review(
        db,
        orders_user_id="merchant-1",
        markers_user_id="main",
        from_date="2026-07-01",
        to_date="2026-08-12",
    ))
    second = _run(build_invoice_review(
        db,
        orders_user_id="merchant-1",
        markers_user_id="main",
        from_date="2026-07-01",
        to_date="2026-08-12",
        page=2,
    ))
    assert len(first["items"]) == 15
    assert first["pages"] == 2
    assert len(second["items"]) == 2


def test_excel_export_refuses_instead_of_silently_truncating(monkeypatch):
    from integrations.qoyod import invoice_review as review_module

    monkeypatch.setattr(review_module, "MAX_EXPORT_ROWS", 1)
    with pytest.raises(ValueError, match="invoice_export_exceeds_1_rows"):
        _run(build_invoice_review(
            _review_db(),
            orders_user_id="merchant-1",
            markers_user_id="main",
            from_date="2026-07-01",
            to_date="2026-08-12",
            include_all=True,
        ))


def test_excel_is_rtl_and_contains_summary_due_date_and_currency():
    report = _run(build_invoice_review(
        _review_db(),
        orders_user_id="merchant-1",
        markers_user_id="main",
        from_date="2026-07-01",
        to_date="2026-08-12",
        include_all=True,
    ))
    workbook = load_workbook(BytesIO(build_invoice_review_workbook(report)))
    assert workbook.sheetnames == ["الملخص", "فواتير قيود"]
    assert workbook["الملخص"].sheet_view.rightToLeft is True
    invoice_sheet = workbook["فواتير قيود"]
    assert invoice_sheet.sheet_view.rightToLeft is True
    headers = [invoice_sheet.cell(2, i).value for i in range(1, 17)]
    assert "تاريخ الاستحقاق" in headers
    assert "العملة" in headers
    assert invoice_sheet.cell(3, 6).value == "2026-08-20"
    assert invoice_sheet.cell(3, 7).value == "SAR"
    assert workbook["الملخص"]["B7"].value == 4  # invoices in range
    assert workbook["الملخص"]["B8"].value == 4  # distinct references
    assert workbook["الملخص"]["B9"].value == 2  # eligible matched
    assert workbook["الملخص"]["B10"].value == 1  # eligible missing
    assert workbook["الملخص"]["B11"].value == 3  # outside eligible


def test_excel_formula_injection_is_neutralised_without_touching_numbers_dates():
    assert _excel_safe_value("=2+2") == "'=2+2"
    assert _excel_safe_value(" +SUM(A1:A2)") == "' +SUM(A1:A2)"
    assert _excel_safe_value("-10+20") == "'-10+20"
    assert _excel_safe_value("@cmd") == "'@cmd"
    assert _excel_safe_value(170.83) == 170.83
    assert _excel_safe_value(date(2026, 8, 12)) == date(2026, 8, 12)

    report = _run(build_invoice_review(
        _review_db(),
        orders_user_id="merchant-1",
        markers_user_id="main",
        from_date="2026-07-01",
        to_date="2026-08-12",
        include_all=True,
    ))
    report["items"][0].update({
        "invoice_number": "=1+1",
        "reference": "+277000001",
        "customer_name": "@customer",
        "status": "-danger",
    })
    workbook = load_workbook(BytesIO(build_invoice_review_workbook(report)))
    row = workbook["فواتير قيود"][3]
    assert row[1].value == "'=1+1"
    assert row[2].value == "'+277000001"
    assert row[3].value == "'@customer"
    assert row[10].value == "'-danger"
    assert row[4].value == "2026-08-12"  # date text unchanged
    assert row[7].value == 100.0          # numeric unchanged


def test_sync_delegates_to_existing_get_only_mirror_sync(monkeypatch):
    from integrations.qoyod import qoyod_invoices_sync as sync_module

    sentinel_client = object()
    calls = []

    async def fake_existing_sync(
        db, *, user_id, api_client, from_date=None, **_kwargs,
    ):
        calls.append((user_id, api_client, from_date))
        return {
            "ok": True, "fetched": 2, "in_scope": 2,
            "created": 0, "updated": 2,
        }

    monkeypatch.setattr(sync_module, "sync_qoyod_invoices",
                        fake_existing_sync)
    result = _run(sync_invoice_review(
        _review_db(),
        orders_user_id="merchant-1",
        markers_user_id="main",
        api_client=sentinel_client,
        from_date="2026-07-01",
        to_date="2026-08-12",
    ))
    assert calls == [("main", sentinel_client, date(2026, 7, 1))]
    assert result["sync_summary"]["ran"] is True
    assert result["sync_summary"]["ok"] is True
    assert result["sync_summary"]["in_scope"] == 2


@pytest.mark.asyncio
async def test_http_contract_and_export():
    db = _review_db()

    async def current_user():
        return {
            "id": "employee-1",
            "role": "operations",
            "created_by": "merchant-1",
        }

    async def fake_key(_db, tenant):
        assert tenant == "main"
        return "secret-never-returned"

    async def fake_client(_db, tenant, key):
        assert (tenant, key) == ("main", "secret-never-returned")
        return object()

    router = APIRouter(prefix="/integrations/qoyod")
    attach_invoice_review_routes(
        router,
        db=db,
        current_user=current_user,
        tenant_id=lambda _user: "main",
        orders_owner_id=lambda user: user["created_by"],
        get_api_key=fake_key,
        build_api_client=fake_client,
    )
    app = FastAPI()
    app.include_router(router, prefix="/api")

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test",
    ) as client:
        response = await client.get(
            "/api/integrations/qoyod/invoice-review",
            params={
                "from_date": "2026-07-01",
                "to_date": "2026-08-12",
            },
        )
        export = await client.get(
            "/api/integrations/qoyod/invoice-review/export.xlsx",
            params={
                "from_date": "2026-07-01",
                "to_date": "2026-08-12",
            },
        )

    assert response.status_code == 200
    body = response.json()
    assert set(body["items"][0]) == {
        "qoyod_invoice_id", "invoice_number", "reference",
        "salla_order_number", "customer_name", "issue_date", "due_date",
        "currency", "total", "paid_amount", "remaining", "status",
        "last_sync_at", "exact_reference_match", "salla_status",
        "salla_total",
    }
    assert export.status_code == 200
    assert export.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert load_workbook(BytesIO(export.content)).sheetnames == [
        "الملخص", "فواتير قيود",
    ]
