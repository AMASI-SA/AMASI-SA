"""Iter-158 — Unified settlements overview (Salla + Tamara + Tabby)."""
import os
import uuid
import io
import datetime as dt

import pytest
import requests
import openpyxl


BASE_URL = (
    os.environ.get("REACT_APP_BACKEND_URL")
    or open("/app/frontend/.env").read()
    .split("REACT_APP_BACKEND_URL=")[1].split("\n")[0].strip()
)


@pytest.fixture
def ctx():
    suffix = uuid.uuid4().hex[:8]
    email = f"i158-{suffix}@example.com"
    pwd = "T#158abcD"
    requests.post(f"{BASE_URL}/api/auth/register",
                  json={"email": email, "password": pwd, "name": "I158"}, timeout=10)
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"email": email, "password": pwd}, timeout=10)
    yield {"hdr": {"Authorization": f"Bearer {r.json()['access_token']}"}}


def _upload_salla(ctx):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice # 7777"
    ws.append(["رقم الطلب", "إجمالي الطلب (ر.س)", "طريقة الدفع",
               "الرسوم (ر.س)", "المستحق قبل الضريبة (ر.س)",
               "الضريبة", "المستحق بعد الضريبة (ر.س)"])
    ws.append(["111", 500.0, "مدى", 10.0, 490.0, 1.5, 488.5])
    buf = io.BytesIO()
    wb.save(buf)
    r = requests.post(
        f"{BASE_URL}/api/payment-settlements/upload",
        headers=ctx["hdr"],
        files={"file": ("salla.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"provider_hint": "salla"}, timeout=20,
    )
    assert r.status_code == 200, r.text
    return r.json()["file_id"]


def test_unified_list_empty(ctx):
    today = dt.date.today()
    r = requests.get(
        f"{BASE_URL}/api/payment-settlements/_overview/unified?year={today.year}&month={today.month}",
        headers=ctx["hdr"], timeout=10,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["rows"] == []
    assert body["year"] == today.year
    assert body["month"] == today.month


def test_unified_list_returns_salla_file(ctx):
    _upload_salla(ctx)
    today = dt.date.today()
    r = requests.get(
        f"{BASE_URL}/api/payment-settlements/_overview/unified?year={today.year}&month={today.month}",
        headers=ctx["hdr"], timeout=10,
    )
    body = r.json()
    assert len(body["rows"]) == 1
    row = body["rows"][0]
    assert row["provider"] == "salla"
    assert row["gross"] == 500.0
    assert row["fees"] == 10.0
    assert row["net_to_bank"] == 488.5


def test_unified_list_month_navigation(ctx):
    """Files belong to the upload month; navigating to a prior month
    must return empty."""
    _upload_salla(ctx)
    # Future-month request → empty
    r = requests.get(
        f"{BASE_URL}/api/payment-settlements/_overview/unified?year=2027&month=1",
        headers=ctx["hdr"], timeout=10,
    )
    assert r.status_code == 200
    assert r.json()["rows"] == []


def test_unified_list_rejects_bad_month(ctx):
    r = requests.get(
        f"{BASE_URL}/api/payment-settlements/_overview/unified?year=2026&month=13",
        headers=ctx["hdr"], timeout=10,
    )
    assert r.status_code == 400


def test_export_excel_with_selected_files(ctx):
    fid = _upload_salla(ctx)
    r = requests.post(
        f"{BASE_URL}/api/payment-settlements/_overview/export-excel",
        headers=ctx["hdr"], json={"file_ids": [fid]}, timeout=15,
    )
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers.get("content-type", "")
    # Verify it's a valid xlsx
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws["A1"].value == "المزوّد"
    assert ws["A2"].value == "SALLA"


def test_export_excel_rejects_empty_selection(ctx):
    r = requests.post(
        f"{BASE_URL}/api/payment-settlements/_overview/export-excel",
        headers=ctx["hdr"], json={"file_ids": []}, timeout=10,
    )
    assert r.status_code == 400
