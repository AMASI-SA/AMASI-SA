import io

import pytest
from fastapi import HTTPException
from openpyxl import Workbook

from excel_parser import MAX_SALLA_ROWS, parse_salla_excel
from products_import_routes import (
    MAX_IMPORT_ROWS,
    _parse_categories_xlsx,
    _parse_products_xlsx,
)


def _workbook_bytes(*, last_row: int, header: tuple[str, ...]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for column, value in enumerate(header, start=1):
        sheet.cell(1, column, value)
    sheet.cell(last_row, 1, "sentinel")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_salla_parser_rejects_excessive_sheet_rows_before_materializing():
    payload = _workbook_bytes(
        last_row=MAX_SALLA_ROWS + 16,
        header=("رقم الطلب", "إجمالي الطلب"),
    )

    with pytest.raises(ValueError, match="50,000"):
        parse_salla_excel(payload)


@pytest.mark.parametrize(
    ("parser", "header"),
    [
        (_parse_categories_xlsx, ("التصنيفات", "هل فرعي", "التصنيف الأساسي")),
        (_parse_products_xlsx, ("رقم المنتج", "اسم المنتج", "التصنيف")),
    ],
)
def test_product_import_parsers_reject_excessive_sheet_rows(parser, header):
    payload = _workbook_bytes(
        last_row=MAX_IMPORT_ROWS + 2,
        header=header,
    )

    with pytest.raises(HTTPException) as exc:
        parser(payload)

    assert exc.value.status_code == 413
    assert "50,000" in str(exc.value.detail)
